import calendar
import hashlib
import json
import os
import re
import requests
import threading
import time
from datetime import date, datetime, timezone, timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, request, jsonify, session, redirect, render_template, url_for, g
from google.auth import default as _gauth_default
from google.cloud import bigquery

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(32))

BQ_PROJECT       = "fluency-finance"
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
REFRESH_SECRET   = os.environ.get("REFRESH_SECRET", "")
ALLOWED_DOMAIN   = "fluencyacademy.io"
MASTER_EMAILS    = {"raphael.bride@fluencyacademy.io", "matheus.schafer@fluencyacademy.io"}
# People Ops: visão GERAL de toda a equipe comercial (todos os colaboradores,
# ativos E inativos), SEM PII de folha e SEM poder de edição (read-only).
# Para quem precisa ver times/ranking/trend mas não está no roster de vendas.
PEOPLE_OPS_EMAILS = {"milena.custodio@fluencyacademy.io", "paula@fluencyacademy.io"}
# Gestor por override: enxerga a equipe comercial como um coordenador (todos os TLs,
# suas equipes e os liderados diretos), com "Ver como" e drill — mas NÃO está no roster
# de vendas. Para quem exerce papel de gestão sem ser coordenador cadastrado.
GESTOR_EMAILS = {"liliane.noga@fluencyacademy.io"}
# Aprovador final: visão geral + aceite próprio que dispara o e-mail para People Ops.
APROVADOR_FINAL_EMAIL = "felipe.yani@fluencyacademy.io"

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

# Coordenador comercial único (recebe sign-off e aprova extras de TODOS os vendedores).
COORDENADOR_EMAIL = "fabio.dias@fluencyacademy.io"
# Master/FP&A que recebe cópia dos sign-offs e extras.
FPA_EMAIL = "raphael.bride@fluencyacademy.io"
# Webhook de e-mail (Apps Script web app). Sem URL configurada → envio é no-op (loga e segue).
MAIL_WEBHOOK_URL    = os.environ.get("MAIL_WEBHOOK_URL", "")
MAIL_WEBHOOK_SECRET = os.environ.get("MAIL_WEBHOOK_SECRET", "")
# SMTP do fechamento — remetente oficial de Finanças (caixa real). App Password no secret.
FINANCE_SMTP_USER = os.environ.get("FINANCE_SMTP_USER", "finance@fluencyacademy.io")
FINANCE_SMTP_PASS = os.environ.get("FINANCE_SMTP_PASS", "")   # App Password (16 chars) — só no Cloud Run
FINANCE_SMTP_HOST = os.environ.get("FINANCE_SMTP_HOST", "smtp.gmail.com")
FINANCE_SMTP_PORT = int(os.environ.get("FINANCE_SMTP_PORT", "587"))
# Cópia fixa no fechamento: RH/People (Giulia) + Paula.
SIGNOFF_CC_FIXO = ["giullia@fluencyacademy.io", "paula@fluencyacademy.io"]
# Destinatários do e-mail de confirmação formal de aceite (financeiro).
FINANCE_CONFIRMACAO_EMAILS = ["giullia@fluencyacademy.io", "paula@fluencyacademy.io"]
# Aviso People Ops: quando 100% do roster aceitou (Milena + Thabita + Matheus, override via env var).
PEOPLE_OPS_NOTIFY_EMAILS = os.environ.get(
    "PEOPLE_OPS_NOTIFY_EMAILS",
    "milena.custodio@fluencyacademy.io,thabita.ruppenthal@fluencyacademy.io,matheus.schafer@fluencyacademy.io",
).split(",")
# Bucket GCS para PDFs de fechamento (gerados no aceite, link enviado por e-mail).
GCS_BUCKET = os.environ.get("GCS_FECHAMENTOS_BUCKET", "fluency-finance-fechamentos")
# Markdown de governança (lido em runtime — sem redeploy ao editar o .md).
GOVERNANCA_PATH = Path(os.environ.get(
    "GOVERNANCA_PATH",
    str(Path(__file__).parent / "runbooks" / "governanca-comissionamento.md")
))

# Drive folder "Commission Sales"
DRIVE_FOLDER_ID  = "1ub0lETiv5XFYKWnK-CGUVItDy_UhSKpf"
# Google Sheet with vendor metas (aba "Meta" + aba "TL")
SHEETS_META_ID   = "1DFaBtFSam1PIzESqvccapO_X52zoQWuBQ_scji5yi3Y"
# Drive folder where monthly Target sheets are uploaded ("upload")
UPLOAD_FOLDER_ID = "1gIiiFCXpvbdRjSYM-MmtcN5zA0Q6txJB"

_MONTH_PT_FULL = {
    1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril",
    5: "maio",    6: "junho",     7: "julho", 8: "agosto",
    9: "setembro",10: "outubro",  11: "novembro", 12: "dezembro",
}

# Maps TL name (as written in Target sheet col C) → email
_TL_NAME_TO_EMAIL: dict[str, str] = {
    "matheus":              "matheus.fernandes@fluencyacademy.io",
    "matheus fernandes":    "matheus.fernandes@fluencyacademy.io",
    "tacyana bueno":        "tacyana.bueno@fluencyacademy.io",
    "tacyana":              "tacyana.bueno@fluencyacademy.io",
    "ana pamplona":         "anaclara.pamplona@fluencyacademy.io",
    "anaclara pamplona":    "anaclara.pamplona@fluencyacademy.io",
    "vanessa lopes":        "vanessa.lopes@fluencyacademy.io",
    "vanessa":              "vanessa.lopes@fluencyacademy.io",
    "fabio":                "fabio.dias@fluencyacademy.io",
    "fabio dias":           "fabio.dias@fluencyacademy.io",
}


# Assistentes que entraram no fim do mês: atingimento tratado como 100% (mult=1.0, vlr=OTE).
# Badge "nota_fim_mes" é injetado em todas as respostas de API onde esses colaboradores aparecem.
_NEWCOMER_ASSISTS: dict[str, set[str]] = {}

def _is_newcomer_assist(email: str, mes: str) -> bool:
    """True se o atingimento desta assistente foi ajustado a 100% por entrada no fim do mês."""
    key = mes[:7] + "-01" if len(mes) >= 7 else mes
    s = _NEWCOMER_ASSISTS.get(key)
    return bool(s and email.lower() in s)

# ── Regras de comissão (espelho de pipeline/calc_comissao.py) ─────────────────
# Taxas por modalidade de pagamento (pré-multiplicador), por modelo de cargo.
_RATES_BY_MODELO: dict[str, dict[str, float]] = {
    "analista":   {"a vista": 0.10, "parcelado": 0.04,  "inteligente": 0.013},
    "deluchi":    {"a vista": 0.08, "parcelado": 0.04,  "inteligente": 0.0225},  # Recuperação: tabela própria
    "tl_vanessa": {"a vista": 0.04, "parcelado": 0.015, "inteligente": 0.005},
    "tl_tacyana": {"a vista": 0.03, "parcelado": 0.01,  "inteligente": 0.005},
}
# OTE (R$) para modelos baseados em atingimento × OTE × mult.
_OTE_BY_MODELO: dict[str, float] = {"assistente": 4000.0, "tl_novo": 7000.0}

def _mult_for_modelo(modelo: str, at: float) -> float | None:
    """Multiplicador pela tabela do modelo. None = modelo sem escada definida (coord)."""
    if modelo == "analista":
        if at < 0.75: return 0.3
        if at < 0.98: return 0.5
        if at < 1.20: return 1.0
        if at < 1.30: return 1.2
        if at < 1.50: return 1.3
        return 1.5
    if modelo == "deluchi":   # Recuperação: tabela própria (máx 1,3×, limiares distintos)
        if at < 0.80: return 0.5
        if at < 0.95: return 0.8
        if at < 1.20: return 1.0
        if at < 1.30: return 1.2
        return 1.3
    if modelo in ("assistente", "tl_novo"):
        if at < 0.70: return 0.3
        if at < 0.80: return 0.5
        if at < 0.90: return 0.7
        if at < 1.10: return 1.0
        return 1.3
    if modelo == "tl_vanessa":
        if at < 0.80: return 0.6
        if at < 0.95: return 0.8
        if at < 1.20: return 1.0
        return 1.2
    if modelo == "tl_tacyana":
        if at < 0.80: return 0.4
        if at < 1.00: return 0.6
        if at < 1.20: return 0.7
        return 1.0
    return None

_MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
_PAID_STATUSES    = {"aprovado", "completo", "approved", "complete"}
_REFUND_STATUSES  = {"reembolsado", "refunded", "reembolso"}
_CHARGE_STATUSES  = {"reclamado", "chargeback"}

# Initialize BQ client with Drive scope so that vw_comissao can read the
# meta_vendedores Google Sheets external table from Cloud Run.
_bq_creds, _ = _gauth_default(scopes=[
    "https://www.googleapis.com/auth/cloud-platform",
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
])
bq = bigquery.Client(project=BQ_PROJECT, credentials=_bq_creds)

# ── Active employees (Supabase) ───────────────────────────────────────────────

_active_cache: tuple[float, set[str]] | None = None

def _get_active_emails() -> set[str] | None:
    """Fetches active employee emails from Supabase with 1h in-memory cache.
    Returns None on failure so callers skip the filter (graceful degradation)."""
    global _active_cache
    if not SUPABASE_URL or not SUPABASE_KEY:
        return None
    now = time.time()
    if _active_cache is not None:
        ts, emails = _active_cache
        if now - ts < 3600:
            return emails
    try:
        today = date.today().isoformat()
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/employees",
            headers={
                "apikey":        SUPABASE_KEY,
                "Authorization": f"Bearer {SUPABASE_KEY}",
            },
            params={
                "select": "email",
                "or":     f"(dismissal_date.is.null,dismissal_date.gt.{today})",
            },
            timeout=5,
        )
        if resp.status_code == 200:
            emails = {r["email"].lower() for r in resp.json() if r.get("email")}
            if emails:
                _active_cache = (now, emails)
                return emails
    except Exception:
        pass
    return None

def _filter_active(rows: list[dict], key: str = "vendedor") -> list[dict]:
    """Remove do resultado APENAS quem está comprovadamente inativo no mês selecionado
    (admitido depois / desligado antes — datas de RH). ⚠️ Blacklist, não whitelist:
    PJ/sem cadastro no `employees` (ex.: ana.pamplona TL, fabio) **permanecem**.
    People Ops vê tudo → não filtra. No-op se o Supabase estiver indisponível."""
    try:
        if _get_role_data().get("role") == "people_ops":
            return rows
    except Exception:
        pass
    inactive = _inactive_emails_for_month(resolve_month())
    if not inactive:
        return rows
    return [r for r in rows if str(r.get(key, "")).lower() not in inactive]

_active_month_cache: dict[str, tuple[float, set[str]]] = {}

def _inactive_emails_for_month(mes: str) -> set[str] | None:
    """E-mails que NÃO estavam ativos no mês `mes` (YYYY-MM-01 ou YYYY-MM), pelas datas de RH:
    admitido DEPOIS do fim do mês OU desligado ANTES do início do mês.
    ⚠️ Blacklist (não whitelist): só entra quem está no `employees` E comprovadamente fora do
    mês. Quem NÃO está no `employees` (PJ — ex.: ana.pamplona, fabio) **não entra** → permanece
    VISÍVEL no filtro. hiring_date nulo = já admitido. Cache 1h por mês.
    None em falha → caller não filtra (degradação graciosa)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return None
    now = time.time()
    hit = _active_month_cache.get(mes)
    if hit and now - hit[0] < 3600:
        return hit[1]
    try:
        y, m = int(mes[:4]), int(mes[5:7])
        first = f"{y:04d}-{m:02d}-01"
        last = (date(y + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1)
                - timedelta(days=1)).isoformat()
        resp = requests.get(
            f"{SUPABASE_URL}/rest/v1/employees",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
            params={"select": "email,hiring_date,dismissal_date"},
            timeout=5,
        )
        if resp.status_code == 200:
            inactive: set[str] = set()
            for r in resp.json():
                em = (r.get("email") or "").lower()
                if not em:
                    continue
                hd, dd = r.get("hiring_date"), r.get("dismissal_date")
                admitido_depois = bool(hd and hd[:10] > last)
                desligado_antes = bool(dd and dd[:10] < first)
                if admitido_depois or desligado_antes:
                    inactive.add(em)
            # cacheia sempre (inclusive set vazio = ninguém inativo)
            _active_month_cache[mes] = (now, inactive)
            return inactive
    except Exception:
        pass
    return None

# ── Auth helpers ──────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "email" not in session:
            qs = request.query_string.decode()
            next_path = request.path + ("?" + qs if qs else "")
            return redirect(url_for("login", next=next_path))
        return f(*args, **kwargs)
    return decorated

def verify_google_token(token: str) -> dict | None:
    resp = requests.get(
        "https://oauth2.googleapis.com/tokeninfo",
        params={"id_token": token},
        timeout=5,
    )
    if resp.status_code != 200:
        return None
    data = resp.json()
    if data.get("aud") != GOOGLE_CLIENT_ID:
        return None
    email = data.get("email", "")
    if not email.endswith(f"@{ALLOWED_DOMAIN}"):
        return None
    return data

def _real_is_master() -> bool:
    """Master de verdade pelo e-mail logado (autoriza o seletor 'Ver como')."""
    return session.get("email", "").lower() in {e.lower() for e in MASTER_EMAILS}

def is_master() -> bool:
    """Master EFETIVO. Um master 'vendo como' outro papel deixa de ser master
    para fins de PII/UI — assim a simulação é fiel e os gates continuam corretos."""
    return _get_role_data()["role"] == "master"

# Papéis que enxergam a visão GERAL (todos os colaboradores): master + People Ops + aprovador.
READ_ALL_ROLES = ("master", "people_ops", "aprovador")
# Papéis read-only (sem poder de edição/aprovação).
READONLY_ROLES = ("people_ops",)
# "Ver como" — quais papéis cada papel REAL pode assumir (sempre de MENOR privilégio):
#   master → People Ops / Gestor / TL / Vendedor (qualquer pessoa)
#   gestor → TL / Vendedor, mas só dentro da própria hierarquia (validado por escopo)
VIEW_AS_BY_ROLE = {
    "master": ("people_ops", "gestor", "aprovador", "tl", "vendedor"),
    "gestor": ("tl", "vendedor"),
}
# Modos "ver como pessoa" (precisam de um e-mail alvo via ?vendedor=).
VIEW_AS_PERSON = ("tl", "vendedor")

def _view_as() -> str | None:
    """Papel que o usuário real está 'vendo como' via ?as=. None se não se aplica/permitido."""
    a = request.args.get("as", "").strip().lower()
    if not a:
        return None
    allowed = VIEW_AS_BY_ROLE.get(_real_role(), ())
    return a if a in allowed else None

def _in_gestor_scope(target: str) -> bool:
    """True se o alvo pertence à hierarquia do gestor real (TLs ou seus liderados).
    Para master (ou não-gestor) sempre True — pode ver qualquer pessoa."""
    rd = _real_role_data()
    if rd["role"] != "gestor":
        return True
    allowed = set(rd.get("reports", []))
    for members in rd.get("tl_reports", {}).values():
        allowed.update(members)
    allowed.update(rd.get("direct_reports", []))
    return target.lower() in allowed

def effective_email() -> str:
    """Identidade efetiva: o e-mail alvo quando se vê 'como TL/Vendedor', senão o logado."""
    if _view_as() in VIEW_AS_PERSON:
        v = request.args.get("vendedor", "").strip().lower()
        if v and v != "todos":
            return v
    return session.get("email", "").lower()

# ── Role resolution ───────────────────────────────────────────────────────────

def _load_vmap() -> dict:
    """Roster do mês mais recente: {email: {cargo, gestor}}."""
    rows = run_query("""
        SELECT
          LOWER(email_vendedor)      AS email,
          LOWER(COALESCE(cargo,''))  AS cargo,
          LOWER(COALESCE(gestor,'')) AS gestor
        FROM `fluency-finance.commission.hierarquia_comercial`
        WHERE mes_venda = (
          SELECT MAX(mes_venda)
          FROM `fluency-finance.commission.hierarquia_comercial`
        )
    """, cache_ttl=300)
    return {r["email"]: {"cargo": r["cargo"], "gestor": r["gestor"]} for r in rows}

def _overview_role_data(vmap: dict, role_name: str) -> dict:
    """Papel de visão GERAL (people_ops/gestor): enxerga TODOS os TLs e suas equipes."""
    all_tls = [e for e, info in vmap.items() if "team leader" in info["cargo"]]
    tlset = set(all_tls)
    tl_reps = {tl: [e for e, info in vmap.items()
                    if info["gestor"] == tl and "team leader" not in info["cargo"]]
               for tl in all_tls}
    # ICs que reportam DIRETO a um coordenador (não a um TL) — não caem em tl_reps.
    direct = [e for e, info in vmap.items()
              if "team leader" not in info["cargo"]
              and info["gestor"] and info["gestor"] not in tlset]
    return {"role": role_name, "reports": all_tls, "tl_reports": tl_reps,
            "direct_reports": direct}

def _compute_role_data(email: str) -> dict:
    """
    Returns dict:
      role     : 'master' | 'gestor' | 'tl' | 'vendedor'
      reports  : [TL emails] for gestor, [vendor emails] for TL, [] otherwise
      tl_reports: {tl_email: [member_emails]} for gestor only
    """
    email = email.lower()
    if email in {e.lower() for e in MASTER_EMAILS}:
        return {"role": "master", "reports": [], "tl_reports": {}}
    try:
        vmap = _load_vmap()
        # Aprovador final: visão geral + pode dar aceite próprio que dispara e-mail People Ops.
        if email == APROVADOR_FINAL_EMAIL.lower():
            return _overview_role_data(vmap, "aprovador")
        # People Ops: visão geral de TODA a equipe comercial, read-only, sem PII.
        # Verificado ANTES do fallback de roster pois não vendem.
        if email in {e.lower() for e in PEOPLE_OPS_EMAILS}:
            return _overview_role_data(vmap, "people_ops")
        # Gestor por override (não está no roster): visão de coordenador sobre toda a equipe.
        if email in {e.lower() for e in GESTOR_EMAILS}:
            return _overview_role_data(vmap, "gestor")
        if email not in vmap:
            return {"role": "vendedor", "reports": [], "tl_reports": {}}
        cargo = vmap[email]["cargo"]
        # Gestor: manages Team Leaders
        tls_under_me = [e for e, info in vmap.items()
                        if info["gestor"] == email and "team leader" in info["cargo"]]
        if tls_under_me:
            tl_reps = {tl: [e for e, info in vmap.items()
                             if info["gestor"] == tl and "team leader" not in info["cargo"]]
                       for tl in tls_under_me}
            # liderados que reportam DIRETO ao gestor (não passam por um TL)
            direct = [e for e, info in vmap.items()
                      if info["gestor"] == email and "team leader" not in info["cargo"]]
            return {"role": "gestor", "reports": tls_under_me,
                    "tl_reports": tl_reps, "direct_reports": direct}
        # TL: has direct vendor reports
        if "team leader" in cargo:
            direct = [e for e, info in vmap.items()
                      if info["gestor"] == email and "team leader" not in info["cargo"]]
            return {"role": "tl", "reports": direct, "tl_reports": {}}
        return {"role": "vendedor", "reports": [], "tl_reports": {}}
    except Exception:
        return {"role": "vendedor", "reports": [], "tl_reports": {}}


def _real_role_data() -> dict:
    """Papel REAL do usuário logado (ignora ?as=). Cacheado por request."""
    if not hasattr(g, "real_role_data"):
        g.real_role_data = _compute_role_data(session.get("email", "").lower())
    return g.real_role_data

def _real_role() -> str:
    return _real_role_data()["role"]

def _get_role_data() -> dict:
    """Role data cached per request via flask.g.
    Aplica o 'Ver como' (?as=) quando um master/gestor real simula outro papel."""
    if not hasattr(g, "role_data"):
        va = _view_as()
        if va in ("people_ops", "gestor", "aprovador"):
            try:
                g.role_data = _overview_role_data(_load_vmap(), va)
            except Exception:
                g.role_data = {"role": va, "reports": [], "tl_reports": {}}
        elif va in VIEW_AS_PERSON:
            # vê como o TL/Vendedor selecionado (papel real daquele e-mail),
            # desde que dentro do escopo (gestor) — senão cai na própria visão real.
            target = effective_email()
            g.role_data = _compute_role_data(target) if _in_gestor_scope(target) else _real_role_data()
        else:
            g.role_data = _real_role_data()
    return g.role_data


def resolve_target(default_email: str) -> str:
    """Returns the target vendedor — honours ?vendedor= param for masters and gestores."""
    role_data = _get_role_data()
    role = role_data["role"]
    if role in READ_ALL_ROLES:
        v = request.args.get("vendedor", "").strip()
        if v and v != "todos":
            return v
    elif role == "gestor":
        v = request.args.get("vendedor", "").strip().lower()
        if v and v != "todos":
            # Only allow TLs or their direct reports
            allowed: set[str] = set(role_data.get("reports", []))
            for members in role_data.get("tl_reports", {}).values():
                allowed.update(members)
            allowed.update(role_data.get("direct_reports", []))
            if v in allowed:
                return v
    elif role == "tl":
        v = request.args.get("vendedor", "").strip().lower()
        # TL só pode drilar membros do PRÓPRIO time (hierarquia)
        if v and v != "todos" and v in {e.lower() for e in role_data.get("reports", [])}:
            return v
    return default_email

def current_month_brt() -> str:
    """Returns YYYY-MM-01 for current month in BRT (UTC-3, Brazil standard)."""
    brt = datetime.now(timezone.utc) - timedelta(hours=3)
    return brt.strftime("%Y-%m-01")

def require_refresh_secret(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not REFRESH_SECRET or request.headers.get("X-Refresh-Secret") != REFRESH_SECRET:
            return jsonify({"error": "unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

def resolve_month() -> str:
    """Returns YYYY-MM-01 for the selected month, defaulting to current month."""
    m = request.args.get("mes", "").strip()
    if m and len(m) == 7:   # expects YYYY-MM
        return m + "-01"
    return current_month_brt()

# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if "email" not in session:
        return redirect(url_for("login"))
    return redirect(url_for("dashboard"))

@app.route("/login")
def login():
    if "email" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html", client_id=GOOGLE_CLIENT_ID)

@app.route("/dashboard")
@login_required
def dashboard():
    role_data = _get_role_data()
    _log_access(role_data["role"])
    return render_template("dashboard.html",
                           email=session["email"],
                           client_id=GOOGLE_CLIENT_ID,
                           is_master=is_master(),
                           real_master=_real_is_master(),
                           role=role_data["role"])

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── Auth endpoint ─────────────────────────────────────────────────────────────

@app.route("/auth/google", methods=["POST"])
def auth_google():
    token = request.json.get("credential") if request.is_json else request.form.get("credential")
    if not token:
        return jsonify({"error": "token ausente"}), 400
    user = verify_google_token(token)
    if not user:
        return jsonify({"error": "token inválido ou domínio não autorizado"}), 401
    session["email"]   = user["email"]
    session["name"]    = user.get("name", user["email"])
    session["picture"] = user.get("picture", "")
    return jsonify({"ok": True, "email": user["email"]})

# ── BQ helpers ────────────────────────────────────────────────────────────────

_QCACHE: dict = {}
_QLOCK = threading.Lock()

def run_query(sql: str, params: list | None = None, cache_ttl: int = 0) -> list[dict]:
    """Executa query no BQ. Se cache_ttl>0, cacheia o resultado em memória por TTL segundos
    (por instância) — usado nas leituras agregadas/globais que 40+ usuários repetem na rajada.
    NÃO usar em dados sensíveis a escrita (aprovações)."""
    if cache_ttl > 0:
        key = hashlib.sha256(
            (sql + "|" + repr([(p.name, p.value) for p in (params or [])])).encode()
        ).hexdigest()
        now = time.time()
        with _QLOCK:
            hit = _QCACHE.get(key)
            if hit and hit[0] > now:
                return hit[1]
        result = _run_query_raw(sql, params)
        with _QLOCK:
            _QCACHE[key] = (now + cache_ttl, result)
            if len(_QCACHE) > 500:   # poda simples p/ não crescer indefinidamente
                for k in [k for k, v in _QCACHE.items() if v[0] <= now]:
                    _QCACHE.pop(k, None)
        return result
    return _run_query_raw(sql, params)

def _run_query_raw(sql: str, params: list | None = None) -> list[dict]:
    job_config = bigquery.QueryJobConfig(query_parameters=params or [])
    rows = bq.query(sql, job_config=job_config).result()
    return [dict(r) for r in rows]

def _log_access(role: str) -> None:
    """Registra um acesso ao dashboard em commission.access_log. Best-effort: nunca trava a página."""
    try:
        ip = (request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
              or request.remote_addr or "")
        bq.insert_rows_json("fluency-finance.commission.access_log", [{
            "email":       session.get("email", ""),
            "accessed_at": datetime.now(timezone.utc).isoformat(),
            "role":        role,
            "ip":          ip,
            "user_agent":  request.headers.get("User-Agent", "")[:500],
            "path":        request.path,
            "name":        session.get("name", ""),
            "picture":     session.get("picture", ""),
        }])
    except Exception:
        pass

# ── E-mail (Apps Script webhook) ───────────────────────────────────────────────

def send_mail(to: list, subject: str, html: str, cc: list | None = None) -> bool:
    """Dispara e-mail via Apps Script web app. Best-effort: nunca derruba a request.
    Não envia em modo 'ver como' (simulação) nem sem webhook configurado."""
    if _view_as():
        app.logger.info("send_mail suprimido (modo ver-como): %s", subject)
        return False
    to = [e for e in dict.fromkeys([(x or "").strip().lower() for x in (to or [])]) if e]
    cc = [e for e in dict.fromkeys([(x or "").strip().lower() for x in (cc or [])]) if e and e not in to]
    if not to:
        return False
    if not MAIL_WEBHOOK_URL:
        app.logger.warning("MAIL_WEBHOOK_URL ausente — e-mail '%s' p/ %s NÃO enviado", subject, to)
        return False
    to_str = ",".join(to)
    cc_str = ",".join(cc) if cc else ""
    try:
        r = requests.post(MAIL_WEBHOOK_URL, json={
            "secret": MAIL_WEBHOOK_SECRET, "to": to_str, "cc": cc_str,
            "subject": subject, "html": html,
        }, timeout=10)
        ok = r.status_code == 200
        if ok:
            app.logger.info("E-mail enviado via webhook [%s]: para=%s cc=%s", subject[:60], to_str, cc_str)
        else:
            app.logger.warning("Webhook de e-mail respondeu %s (body=%s): to=%s subject=%s",
                               r.status_code, r.text[:200], to_str, subject)
        return ok
    except Exception as e:
        app.logger.warning("Falha ao enviar e-mail '%s': %s", subject, e)
        return False

_GCS_SA_EMAIL = "commission-dashboard@fluency-finance.iam.gserviceaccount.com"

def _upload_pdf_gcs(pdf_bytes: bytes, vend: str, mes: str) -> str | None:
    """Faz upload do PDF para GCS e retorna URL assinada válida por 7 dias. None se falhar."""
    try:
        import uuid
        import google.auth
        import google.auth.transport.requests as google_auth_requests
        from google.cloud import storage as gcs
        from datetime import timedelta
        credentials, _ = google.auth.default(
            scopes=["https://www.googleapis.com/auth/cloud-platform"]
        )
        credentials.refresh(google_auth_requests.Request())
        client = gcs.Client(credentials=credentials)
        bucket = client.bucket(GCS_BUCKET)
        blob_name = f"fechamentos/{mes[:7]}/{vend.split('@')[0]}_{uuid.uuid4().hex[:8]}.pdf"
        blob = bucket.blob(blob_name)
        blob.upload_from_string(pdf_bytes, content_type="application/pdf")
        url = blob.generate_signed_url(
            expiration=timedelta(days=7),
            method="GET",
            version="v4",
            service_account_email=_GCS_SA_EMAIL,
            access_token=credentials.token,
        )
        return url
    except Exception as e:
        app.logger.warning("GCS upload falhou: %s", e)
        return None


def _brl(v) -> str:
    """Formata número como R$ no padrão pt-BR (1.234,56)."""
    s = f"{float(v or 0):,.2f}"
    return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")

def send_mail_smtp(to: list, subject: str, html: str, cc: list | None = None,
                   attachments: list | None = None) -> bool:
    """Envia e-mail autenticado pela caixa de Finanças (finance@) via SMTP.
    attachments = [(filename, bytes, mimetype)]. Best-effort: nunca derruba a request.
    Não envia em 'ver como' nem sem App Password configurado."""
    if _view_as():
        app.logger.info("send_mail_smtp suprimido (modo ver-como): %s", subject)
        return False
    to = [e for e in dict.fromkeys([(x or "").strip().lower() for x in (to or [])]) if e]
    cc = [e for e in dict.fromkeys([(x or "").strip().lower() for x in (cc or [])]) if e and e not in to]
    if not to:
        return False
    if not FINANCE_SMTP_PASS:
        app.logger.warning("FINANCE_SMTP_PASS ausente — e-mail '%s' p/ %s NÃO enviado (configure o App Password)", subject, to)
        return False
    import smtplib
    from email.message import EmailMessage
    msg = EmailMessage()
    msg["From"] = FINANCE_SMTP_USER
    msg["To"]   = ", ".join(to)
    if cc:
        msg["Cc"] = ", ".join(cc)
    msg["Subject"] = subject
    msg.set_content("Este e-mail requer um cliente compatível com HTML.")
    msg.add_alternative(html, subtype="html")
    for fn, data, mime in (attachments or []):
        maintype, _, subtype = (mime or "application/octet-stream").partition("/")
        msg.add_attachment(data, maintype=maintype, subtype=(subtype or "octet-stream"), filename=fn)
    try:
        with smtplib.SMTP(FINANCE_SMTP_HOST, FINANCE_SMTP_PORT, timeout=20) as s:
            s.starttls()
            s.login(FINANCE_SMTP_USER, FINANCE_SMTP_PASS)
            s.send_message(msg, to_addrs=to + cc)
        app.logger.info("E-mail enviado (finance@) '%s' p/ %s cc %s", subject, to, cc)
        return True
    except Exception as e:
        app.logger.warning("Falha SMTP '%s': %s", subject, e)
        return False

def _build_signoff_pdf(nome, email, mes_label, cargo, tl, summary, txs) -> bytes:
    """Gera o PDF anexo do fechamento: cabeçalho + resumo + tabela completa de transações."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title=f"Fechamento comissão {nome} {mes_label}")
    ss = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=ss["Title"], fontSize=15, spaceAfter=2, textColor=colors.HexColor("#7B5CF6"))
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"))
    cell = ParagraphStyle("cell", parent=ss["Normal"], fontSize=7.5, leading=9)
    PURPLE = colors.HexColor("#7B5CF6")
    el = []
    el.append(Paragraph("Fluency · Fechamento de Comissão", h))
    el.append(Paragraph(f"<b>{nome}</b> &lt;{email}&gt; · {cargo or '—'} · TL: {tl or '—'} · Competência: <b>{mes_label}</b>", sub))
    el.append(Spacer(1, 8))
    # Resumo — espelha os KPIs do card do vendedor no dashboard
    vm = summary.get("valor_meta")
    n_churn = summary.get("churns", 0)
    churn_label = f"{_brl(summary.get('churn'))} ({n_churn} tx)" if n_churn else _brl(summary.get("churn"))
    res = [
        ["GBV bruto",          _brl(summary.get("gbv_bruto")),
         "Meta",               _brl(vm) if vm else "—"],
        ["Churn",              churn_label,
         "Atingimento",        f"{(summary.get('ating') or 0)*100:.2f}%".replace(".", ",")],
        ["GBV líquido",        _brl(summary.get("gbv_liq")),
         "Multiplicador",      f"{(summary.get('mult') or 0):.1f}×".replace(".", ",")],
        ["Comissão (antes mult)", _brl(summary.get("total_comissao")),
         "Comissão final",     _brl(summary.get("vlr_final"))],
    ]
    rt = Table(res, colWidths=[38*mm, 45*mm, 38*mm, 45*mm])
    rt.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 8.5),
        ("TEXTCOLOR", (0,0), (0,-1), colors.HexColor("#64748b")),
        ("TEXTCOLOR", (2,0), (2,-1), colors.HexColor("#64748b")),
        ("FONTNAME", (1,0), (1,-1), "Helvetica-Bold"),
        ("FONTNAME", (3,0), (3,-1), "Helvetica-Bold"),
        ("FONTSIZE", (3,3), (3,3), 10),
        ("TEXTCOLOR", (3,3), (3,3), PURPLE),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4), ("TOPPADDING", (0,0), (-1,-1), 4),
        ("LINEBELOW", (0,0), (-1,-2), 0.3, colors.HexColor("#e2e8f0")),
    ]))
    el.append(rt)
    el.append(Spacer(1, 10))
    el.append(Paragraph("Detalhamento das transações", ParagraphStyle("h2", parent=ss["Heading2"], fontSize=10)))
    el.append(Spacer(1, 4))
    # Tabela de transações — mesmas colunas da visão do vendedor no dashboard
    head = ["Transaction ID", "Data", "Cliente", "GBV", "Parcela", "Forma", "GBV Líq.", "Comissão", "Status", "Origem"]
    data = [head]
    for t in txs:
        churn_sfx = " *" if t.get("is_churn") else ""
        liq_cell = Paragraph(_brl(t.get("gbv_liquido")) + churn_sfx, cell)
        data.append([
            Paragraph(str(t.get("transaction_id") or "—"), cell),
            str(t.get("data_contrato") or "—"),
            Paragraph(str(t.get("cliente_email") or "—"), cell),
            _brl(t.get("gbv")), _brl(t.get("parcela")),
            (t.get("forma_pagamento") or "—"),
            liq_cell,
            (_brl(t.get("comissao")) if t.get("comissao") is not None else "—"),
            str(t.get("transaction_status") or "—"),
            str(t.get("origem") or "—"),
        ])
    tbl = Table(data, repeatRows=1, colWidths=[28*mm,17*mm,44*mm,22*mm,22*mm,18*mm,22*mm,22*mm,17*mm,15*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), PURPLE),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 7.2),
        ("ALIGN", (3,1), (7,-1), "RIGHT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f7f8fb")]),
        ("LINEBELOW", (0,0), (-1,-1), 0.25, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0,0), (-1,-1), 3), ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    el.append(tbl)
    el.append(Spacer(1, 10))
    el.append(Paragraph(
        f"Confirmado por <b>{nome}</b> ({email}) em {summary.get('signed_at','')} · "
        f"Documento gerado automaticamente pelo dashboard de Comissões — Fluency.", sub))
    doc.build(el)
    return buf.getvalue()

def _build_signoff_xls(nome, email, mes_label, cargo, tl, summary, txs) -> bytes:
    """Gera o XLSX do fechamento: resumo de KPIs + tabela de transações (mesma estrutura do PDF)."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    PURPLE     = "7B5CF6"
    LIGHT_GRAY = "F7F8FB"
    LABEL_CLR  = "64748B"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fechamento"
    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = "Fluency · Fechamento de Comissão"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=PURPLE)
    ws.merge_cells("A2:J2")
    ws["A2"] = f"{nome} <{email}> · {cargo or '—'} · TL: {tl or '—'} · Competência: {mes_label}"
    ws["A2"].font = Font(name="Calibri", size=9, color=LABEL_CLR)
    # ── Resumo KPIs (linhas 4-7) ───────────────────────────────────────────────
    vm = summary.get("valor_meta")
    n_churn = summary.get("churns", 0)
    churn_lbl = f"{_brl(summary.get('churn'))} ({n_churn} tx)" if n_churn else _brl(summary.get("churn"))
    res = [
        ("GBV bruto",             _brl(summary.get("gbv_bruto")),
         "Meta",                  _brl(vm) if vm else "—"),
        ("Churn",                 churn_lbl,
         "Atingimento",           f"{(summary.get('ating') or 0)*100:.2f}%".replace(".", ",")),
        ("GBV líquido",           _brl(summary.get("gbv_liq")),
         "Multiplicador",         f"{(summary.get('mult') or 0):.1f}×".replace(".", ",")),
        ("Comissão (antes mult)", _brl(summary.get("total_comissao")),
         "Comissão final",        _brl(summary.get("vlr_final"))),
    ]
    for i, (l1, v1, l2, v2) in enumerate(res, start=4):
        ws.cell(i, 1, l1).font = Font(size=9, color=LABEL_CLR)
        ws.cell(i, 2, v1).font = Font(size=9, bold=True)
        ws.cell(i, 3, l2).font = Font(size=9, color=LABEL_CLR)
        is_final = (i == 7)
        c = ws.cell(i, 4, v2)
        c.font = Font(size=10 if is_final else 9, bold=is_final,
                      color=PURPLE if is_final else "000000")
    # ── Seção transações ───────────────────────────────────────────────────────
    ws.merge_cells("A9:J9")
    ws["A9"] = "Detalhamento das transações"
    ws["A9"].font = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor=PURPLE)
    hdr_font = Font(bold=True, color="FFFFFF", size=8)
    headers = ["Transaction ID", "Data", "Cliente", "GBV", "Parcela", "Forma",
               "GBV Líq.", "Comissão", "Status", "Origem"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(10, j, h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for idx, t in enumerate(txs, start=11):
        churn_sfx = " *" if t.get("is_churn") else ""
        row = [
            str(t.get("transaction_id") or "—"),
            str(t.get("data_contrato") or "—"),
            str(t.get("cliente_email") or "—"),
            _brl(t.get("gbv")),
            _brl(t.get("parcela")),
            str(t.get("forma_pagamento") or "—"),
            _brl(t.get("gbv_liquido")) + churn_sfx,
            _brl(t.get("comissao")) if t.get("comissao") is not None else "—",
            str(t.get("transaction_status") or "—"),
            str(t.get("origem") or "—"),
        ]
        bg = PatternFill("solid", fgColor=(LIGHT_GRAY if idx % 2 == 0 else "FFFFFF"))
        for j, val in enumerate(row, 1):
            c = ws.cell(idx, j, val); c.fill = bg; c.font = Font(size=8)
    # ── Larguras de coluna ─────────────────────────────────────────────────────
    for j, w in enumerate([30, 12, 35, 14, 14, 14, 14, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _render_comissao_chart(dados: dict) -> bytes:
    """PNG bytes: barra empilhada GBV líquido+churn + callout comissão. Headless (Agg)."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from io import BytesIO
    bruto    = float(dados.get("gbv_bruto") or 0)
    liquido  = float(dados.get("gbv_liquido") or dados.get("gbv_liq") or 0)
    churn    = float(dados.get("gbv_churn") or dados.get("churn") or 0) or max(bruto - liquido, 0)
    comissao = float(dados.get("comissao") or dados.get("vlr_final") or 0)
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 3.2),
                                   gridspec_kw={"height_ratios": [2, 1]})
    fig.patch.set_facecolor("white")
    ax1.barh([0], [liquido], color="#7B5CF6",
             label=f"Líquido — {_brl(liquido)}" if liquido else "Sem GBV no período")
    if churn > 0:
        ax1.barh([0], [churn], left=[liquido], color="#E91E8C", label=f"Churn — {_brl(churn)}")
    ax1.set_xlim(0, max(bruto, 1) * 1.05)
    ax1.set_yticks([])
    ax1.set_title(f"GBV bruto: {_brl(bruto)}", loc="left", fontsize=11, fontweight="bold", color="#1F1F1F")
    ax1.legend(loc="upper right", frameon=False, fontsize=9)
    for spine in ("top", "right", "left"):
        ax1.spines[spine].set_visible(False)
    ax1.tick_params(left=False, labelsize=8, colors="#666")
    ax2.axis("off")
    ax2.text(0.0, 0.5, "Comissão apurada", fontsize=10, color="#666", va="center")
    ax2.text(1.0, 0.5, _brl(comissao), fontsize=20, color="#7B5CF6",
             fontweight="bold", va="center", ha="right")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=144, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()

def _commission_chain(vendedor_email: str, mes: str) -> dict:
    """Cadeia de aprovação/notificação de um vendedor no mês:
    {vendedor, tl, coord, master}. tl = gestor da hierarquia_comercial (e-mail);
    coord = COORDENADOR_EMAIL; master = FP&A. tl pode coincidir com coord (vendedor direto)."""
    vend = (vendedor_email or "").strip().lower()
    rows = run_query("""
        SELECT LOWER(gestor) AS gestor
        FROM `fluency-finance.commission.hierarquia_comercial`
        WHERE LOWER(email_vendedor) = LOWER(@v) AND mes_venda = DATE(@mes)
        LIMIT 1
    """, [
        bigquery.ScalarQueryParameter("v",   "STRING", vend),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])
    tl = (rows[0]["gestor"] if rows and rows[0].get("gestor") else "") or COORDENADOR_EMAIL
    return {"vendedor": vend, "tl": tl, "coord": COORDENADOR_EMAIL, "master": FPA_EMAIL.lower()}

def _approved_extras(vendedor_email: str, mes: str) -> list[dict]:
    """Extras EFETIVOS (TL e coord aprovaram) de um vendedor no mês — para agregar no dash."""
    return run_query("""
        SELECT id, transaction_id, vendedor, gbv, modality_payment, is_churn, nota, created_at
        FROM `fluency-finance.commission.extras_vendedores`
        WHERE LOWER(vendedor) = LOWER(@v)
          AND competencia = DATE(@mes)
          AND status_tl = 'aprovado' AND status_coord = 'aprovado'
        ORDER BY created_at
    """, [
        bigquery.ScalarQueryParameter("v",   "STRING", vendedor_email),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])

def _approved_extras_bulk(mes: str) -> dict[str, list[dict]]:
    """Todos os extras aprovados (TL+coord) do mês, agrupados por email do vendedor.
    Retorna {email_lower: [{gbv, modality_payment, is_churn}, …]}."""
    rows = run_query("""
        SELECT LOWER(vendedor) AS vendedor, gbv, modality_payment, is_churn
        FROM `fluency-finance.commission.extras_vendedores`
        WHERE competencia = DATE(@mes)
          AND status_tl = 'aprovado' AND status_coord = 'aprovado'
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=90)
    out: dict[str, list] = {}
    for r in rows:
        email = str(r["vendedor"] or "").lower()
        out.setdefault(email, []).append({
            "gbv":              float(r["gbv"] or 0),
            "modality_payment": str(r.get("modality_payment") or ""),
            "is_churn":         int(r.get("is_churn") or 0),
        })
    return out

def _extras_for_display(vendedor_email: str, mes: str) -> list[dict]:
    """TODOS os extras (HP) do vendedor no mês — pendentes/aprovados/rejeitados — com os
    dois status (TL e coord) p/ exibir como linha na aba Transações com a coluna de aprovação."""
    return run_query("""
        SELECT id, transaction_id, vendedor, gbv, modality_payment, is_churn, nota,
               created_at, status_tl, status_coord
        FROM `fluency-finance.commission.extras_vendedores`
        WHERE LOWER(vendedor) = LOWER(@v) AND competencia = DATE(@mes)
        ORDER BY created_at
    """, [
        bigquery.ScalarQueryParameter("v",   "STRING", vendedor_email),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])

# ── API: months ───────────────────────────────────────────────────────────────

@app.route("/api/months")
@login_required
def api_months():
    role  = _get_role_data()["role"]
    vend  = request.args.get("vendedor", "").strip().lower()
    # mes + se o mês tem comissão fechada (vlr_final_comissao > 0) -> default abre no
    # mês fechado mais recente, não no mês corrente ainda vazio.
    # Vínculo Colaborador → Mês: se um colaborador estiver selecionado, retorna só os meses
    # em que ELE tem dados (resolve_target valida o escopo/hierarquia).
    if role in ("master", "gestor", "tl", "people_ops") and (not vend or vend == "todos"):
        sql = """
            SELECT DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) AS mes,
                   SUM(COALESCE(vlr_final_comissao, 0)) AS comissao
            FROM `fluency-finance.commission.vw_comissao`
            WHERE vendedor IS NOT NULL AND vendedor != '#N/A'
            GROUP BY mes ORDER BY mes DESC LIMIT 24
        """
        rows = run_query(sql, cache_ttl=90)
    else:
        target = resolve_target(effective_email())
        sql = """
            SELECT DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) AS mes,
                   SUM(COALESCE(vlr_final_comissao, 0)) AS comissao
            FROM `fluency-finance.commission.vw_comissao`
            WHERE LOWER(vendedor) = LOWER(@email)
              AND vendedor IS NOT NULL AND vendedor != '#N/A'
            GROUP BY mes ORDER BY mes DESC LIMIT 24
        """
        rows = run_query(sql, [bigquery.ScalarQueryParameter("email", "STRING", target)])
    months = [str(r["mes"]) for r in rows]
    com_meses = [str(r["mes"]) for r in rows if float(r["comissao"] or 0) > 0]
    default = com_meses[0] if com_meses else (months[0] if months else None)
    return jsonify({"months": months, "default": default})

# ── API: vendors (master only) ────────────────────────────────────────────────

@app.route("/api/vendors")
@login_required
def api_vendors():
    role_data = _get_role_data()
    role = role_data["role"]
    if role in READ_ALL_ROLES:
        sql = """
            SELECT DISTINCT vendedor
            FROM `fluency-finance.commission.vw_comissao`
            WHERE vendedor IS NOT NULL AND vendedor != '#N/A'
            ORDER BY vendedor
        """
        rows = run_query(sql, cache_ttl=90)
        rows = _filter_active(rows, "vendedor")
        return jsonify([r["vendedor"] for r in rows])
    elif role == "gestor":
        # escopo completo do gestor (liderados via TL + os que reportam direto a ele)
        scope = {m for ms in role_data.get("tl_reports", {}).values() for m in ms}
        scope.update(role_data.get("direct_reports", []))
        inactive = _inactive_emails_for_month(resolve_month())   # quem saiu/ainda não entrou no mês
        def _visivel(lst):
            # mostra TODOS, exceto quem está comprovadamente inativo no mês.
            # PJ/sem cadastro no employees (ex.: ana.pamplona TL) NÃO é filtrado → fica visível.
            return sorted(e for e in lst if not inactive or e.lower() not in inactive)
        tl_filter = request.args.get("tl", "").strip().lower()
        if tl_filter:
            # vínculo TL → Vendedor: só o time do TL selecionado
            return jsonify(_visivel(role_data.get("tl_reports", {}).get(tl_filter, [])))
        if request.args.get("all") == "1":
            # dropdown "Vendedor" (vendedores + assistentes) e "Ver como Vendedor"
            return jsonify(_visivel(scope))
        # dropdown "Team Leader" — todos os TLs (PJ inclusos)
        return jsonify(_visivel(role_data.get("reports", [])))
    elif role == "tl":
        # TL: o PRÓPRIO email primeiro (para auto-drill) + time (liderados diretos)
        inactive = _inactive_emails_for_month(resolve_month())
        team = role_data.get("reports", [])
        own = effective_email().lower()
        members = sorted(e for e in team if not inactive or e.lower() not in inactive)
        return jsonify([own] + [e for e in members if e != own])
    return jsonify({"error": "forbidden"}), 403

# ── API: summary ──────────────────────────────────────────────────────────────

@app.route("/api/summary")
@login_required
def api_summary():
    target  = resolve_target(effective_email())
    mes     = resolve_month()
    sql = """
        SELECT
          vendedor,
          DATE(contract_created_at_brt_timestamp)                   AS competencia,
          CAST(gbv                               AS NUMERIC)        AS gbv,
          CAST(gbv_churn_descontado_transaction  AS NUMERIC)        AS gbv_liquido,
          CAST(gbv_apenas_churn_transaction      AS NUMERIC)        AS gbv_churn,
          CAST(qtd_is_churn_transaction          AS INT64)          AS churns,
          CAST(COALESCE(comissao_inteligente, 0) AS NUMERIC)        AS comissao_inteligente,
          CAST(COALESCE(comissao_parcelado,   0) AS NUMERIC)        AS comissao_parcelado,
          CAST(COALESCE(comissao_a_vista,     0) AS NUMERIC)        AS comissao_a_vista,
          CAST(COALESCE(total_comissao,       0) AS NUMERIC)        AS total_comissao,
          CAST(COALESCE(atingimento_meta,     0) AS NUMERIC)        AS atingimento_meta,
          CAST(COALESCE(multiplicador,        0) AS NUMERIC)        AS multiplicador,
          CAST(COALESCE(vlr_final_comissao,   0) AS NUMERIC)        AS vlr_final_comissao,
          CASE WHEN COALESCE(atingimento_meta, 0) > 0
               THEN CAST(SAFE_DIVIDE(
                 CAST(gbv_churn_descontado_transaction AS FLOAT64),
                 CAST(atingimento_meta AS FLOAT64)
               ) AS NUMERIC)
               ELSE NULL
          END                                                        AS valor_meta,
          CAST(COALESCE(meta_proporcional, 0) AS NUMERIC)            AS meta_proporcional,
          CAST(COALESCE(dias_decorridos,   0) AS INT64)              AS dias_decorridos,
          is_projecao                                                AS is_projecao
        FROM `fluency-finance.commission.vw_comissao`
        WHERE LOWER(vendedor) = LOWER(@email)
          AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
          AND vendedor IS NOT NULL AND vendedor != '#N/A'
        LIMIT 1
    """
    rows = run_query(sql, [
        bigquery.ScalarQueryParameter("email", "STRING", target),
        bigquery.ScalarQueryParameter("mes",   "DATE",   mes),
    ])
    # Extras (HP) aprovados por TL+coord — agregados SÓ no dash (GBV pós-churn).
    extras = _approved_extras(target, mes)
    extras_gbv = sum(float(e["gbv"]) for e in extras
                     if e.get("gbv") is not None and not int(e.get("is_churn") or 0))
    extras_count = len(extras)

    if not rows:
        if extras_count == 0:
            return jsonify(None)
        # Sem linha base no snapshot, mas há extras aprovados → expõe só os extras.
        return jsonify({"vendedor": target, "competencia": mes, "gbv": extras_gbv,
                        "gbv_liquido": extras_gbv, "extras_gbv": extras_gbv,
                        "extras_count": extras_count, "prev": None})
    row = rows[0]
    row["competencia"] = str(row["competencia"]) if row["competencia"] else None
    # NUMERIC do BQ vira Decimal → Flask serializa como STRING; o front chama .toFixed()
    # e quebra (renderSummary morre, trava o "Carregando"). Converte os numéricos da linha
    # principal p/ float (o `prev` já fazia isso; a linha principal não fazia).
    for _k, _v in list(row.items()):
        if _k != "competencia" and not isinstance(_v, (str, bool)) and hasattr(_v, "__float__"):
            row[_k] = float(_v)
    row["valor_meta"]  = float(row["valor_meta"]) if row.get("valor_meta") is not None else None
    # TL: o Resumo reflete o TIME (GBV/meta/atingimento do time); a comissão/multiplicador
    # continuam os do próprio TL (já calculados sobre o time pelo pipeline).
    is_proj = bool(row.get("is_projecao"))
    cargo_t = _load_vmap().get(target.lower(), {}).get("cargo", "")
    if "team leader" in cargo_t:
        team = _build_team_totals(mes, target)
        tg = float(team.get("gbv_total") or 0)
        row["gbv"]              = float(team.get("gbv_bruto") or tg)   # bruto real do time
        row["gbv_liquido"]      = tg                                   # pós-churn (base do atingimento)
        row["gbv_churn"]        = float(team.get("churn") or 0.0)      # churn real do time
        row["churns"]           = 0
        if is_proj:
            # projeção: meta proporcional (meta/30 × dias) já computada na linha do TL
            mp = float(row.get("meta_proporcional") or 0)
            row["valor_meta"]       = mp or None
            row["atingimento_meta"] = (tg / mp) if mp > 0 else float(row.get("atingimento_meta") or 0)
        else:
            tm = team.get("meta_total")
            row["valor_meta"]       = float(tm) if tm is not None else row.get("valor_meta")
            row["atingimento_meta"] = (tg / float(tm)) if tm else float(row.get("atingimento_meta") or 0)
    elif is_proj:
        # projeção (vendedor/assistente): meta proporcional + atingimento já vêm da projecao
        mp = float(row.get("meta_proporcional") or 0)
        if mp > 0:
            row["valor_meta"] = mp
            row["atingimento_meta"] = float(row.get("gbv_liquido") or 0) / mp
    else:
        # Meta vem do roster (hierarquia_comercial). O SAFE_DIVIDE derivado fica NULL quando
        # GBV=0 → escondia a meta de quem ainda não vendeu (assistentes novos). Usa a cadastrada.
        mr = run_query("""
            SELECT CAST(COALESCE(valor_meta,0) AS NUMERIC) AS vm,
                   CAST(COALESCE(ote_fator, 1.0) AS FLOAT64) AS ote_fator
            FROM `fluency-finance.commission.hierarquia_comercial`
            WHERE LOWER(email_vendedor)=LOWER(@email) AND mes_venda=DATE(@mes) LIMIT 1
        """, [bigquery.ScalarQueryParameter("email", "STRING", target),
              bigquery.ScalarQueryParameter("mes",   "DATE",   mes)], cache_ttl=300)
        if mr and mr[0]["vm"] is not None and float(mr[0]["vm"]) > 0:
            rm = float(mr[0]["vm"])
            row["valor_meta"] = rm
            row["atingimento_meta"] = float(row.get("gbv_liquido") or 0) / rm
        row["ote_fator"] = float(mr[0]["ote_fator"]) if mr else 1.0
    # Recalcula vlr_final para modelos OTE com ote_fator proporcional (dias trabalhados no mês).
    _modelo_ote = _modelo_comissao(target, cargo_t)
    row["vlr_final_comissao"] = _corrige_vlr_ote(
        float(row.get("vlr_final_comissao") or 0), _modelo_ote,
        float(row.get("atingimento_meta") or 0), float(row.get("multiplicador") or 0),
        float(row.get("ote_fator") or 1.0))

    # Soma o GBV dos extras aprovados aos totais exibidos (transparente via extras_gbv).
    row["extras_gbv"]   = extras_gbv
    row["extras_count"] = extras_count
    if extras_gbv:
        row["gbv"]         = float(row["gbv"]) + extras_gbv
        row["gbv_liquido"] = float(row["gbv_liquido"]) + extras_gbv

    # Recalcula atingimento, multiplicador e vlr_final quando há extras aprovados.
    # Só para vendedores/assistentes em mês fechado (não TL, não projeção).
    if extras_gbv and not is_proj and "team leader" not in cargo_t:
        vm = float(row.get("valor_meta") or 0)
        new_liq = float(row["gbv_liquido"])
        new_ating = (new_liq / vm) if vm > 0 else float(row.get("atingimento_meta") or 0)
        row["atingimento_meta"] = new_ating
        modelo = _modelo_comissao(target, cargo_t)
        new_mult = _mult_for_modelo(modelo, new_ating)
        if new_mult is not None:
            row["multiplicador"] = new_mult
            if modelo in _RATES_BY_MODELO:
                # Comissão transacional: soma a contribuição dos extras por modalidade
                rates = _RATES_BY_MODELO[modelo]
                extras_com = sum(
                    float(e.get("gbv") or 0) * rates.get(e.get("modality_payment") or "", 0)
                    for e in extras
                )
                new_total = float(row.get("total_comissao") or 0) + extras_com
                row["total_comissao"]      = new_total
                row["vlr_final_comissao"]  = new_total * new_mult
            elif modelo in _OTE_BY_MODELO:
                # OTE-based: atingimento × OTE × mult
                row["vlr_final_comissao"] = new_ating * _OTE_BY_MODELO[modelo] * new_mult

    # Previous month for comparison
    from datetime import date
    mes_date = date.fromisoformat(mes)
    if mes_date.month == 1:
        prev_mes = date(mes_date.year - 1, 12, 1).isoformat()
    else:
        prev_mes = date(mes_date.year, mes_date.month - 1, 1).isoformat()

    prev_rows = run_query(sql, [
        bigquery.ScalarQueryParameter("email", "STRING", target),
        bigquery.ScalarQueryParameter("mes",   "DATE",   prev_mes),
    ])
    prev = None
    if prev_rows:
        prev = prev_rows[0]
        prev["competencia"] = str(prev["competencia"]) if prev["competencia"] else None
        for k, v in prev.items():
            if hasattr(v, '__float__'):
                prev[k] = float(v)
    row["prev"] = prev
    # modelo de comissão (define a escada de multiplicadores correta no front)
    row["modelo"]       = _modelo_comissao(target, cargo_t)
    row["nota_fim_mes"] = _is_newcomer_assist(target, mes)
    return jsonify(row)

def _corrige_vlr_ote(vlr: float, modelo: str, atingimento: float, mult: float, ote_fator: float = 1.0) -> float:
    """Modelos OTE: recalcula atingimento × OTE × ote_fator × mult.
    ote_fator < 1.0 para quem entrou no meio do mês (proporcional por dias trabalhados)."""
    if modelo in _OTE_BY_MODELO:
        return atingimento * _OTE_BY_MODELO[modelo] * ote_fator * mult
    return vlr

def _modelo_comissao(email: str, cargo: str) -> str:
    """Identifica o modelo p/ a escada de multiplicadores no dashboard."""
    e = (email or "").lower(); c = (cargo or "").lower()
    if e == COORDENADOR_EMAIL:                 return "coord"
    if e == "ana.deluchi@fluencyacademy.io":   return "deluchi"   # Recuperação: tabela própria (até 1,3)
    if e == "vanessa.lopes@fluencyacademy.io": return "tl_vanessa"
    if e == "tacyana.bueno@fluencyacademy.io": return "tl_tacyana"
    if "team leader" in c:                     return "tl_novo"   # Ana/Matheus = OTE
    if c == "assistente":                      return "assistente"
    return "analista"

# ── API: ranking ──────────────────────────────────────────────────────────────

def _name_from_email(email: str) -> str:
    return " ".join(p.capitalize() for p in email.split("@")[0].split("."))


def _initial_from_email(email: str) -> str:
    return " ".join(p[0].upper() + "." for p in email.split("@")[0].split("."))


_RANKING_SQL_TEAM = """
    SELECT
      ROW_NUMBER() OVER (ORDER BY COALESCE(h.gbv_churn_descontado_transaction,0) DESC) AS posicao,
      h.vendedor,
      CAST(COALESCE(h.gbv_churn_descontado_transaction,0) AS NUMERIC) AS gbv_liquido,
      CAST(COALESCE(h.atingimento_meta, 0) AS NUMERIC)                AS atingimento_meta,
      CAST(COALESCE(h.multiplicador, 0) AS NUMERIC)                   AS multiplicador,
      CAST(COALESCE(h.vlr_final_comissao, 0) AS NUMERIC)              AS vlr_final_comissao,
      CAST(COALESCE(h.total_comissao, 0) AS NUMERIC)                  AS total_comissao,
      CAST(COALESCE(mv.valor_meta, 0) AS NUMERIC)                     AS valor_meta,
      LOWER(COALESCE(mv.cargo, ''))                                   AS cargo
    FROM `fluency-finance.commission.vw_comissao` h
    JOIN `fluency-finance.commission.hierarquia_comercial` mv
      ON LOWER(mv.email_vendedor) = LOWER(h.vendedor)
      AND mv.mes_venda = DATE(@mes)
      AND LOWER(mv.gestor) = LOWER(@tl_email)
      AND LOWER(COALESCE(mv.cargo,'')) NOT LIKE '%team leader%'  -- só o time; NENHUM TL na lista
    WHERE DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
      AND h.vendedor IS NOT NULL AND h.vendedor != '#N/A'
    ORDER BY gbv_liquido DESC
"""

_RANKING_SQL_ALL = """
    SELECT
      ROW_NUMBER() OVER (ORDER BY COALESCE(h.gbv_churn_descontado_transaction,0) DESC) AS posicao,
      h.vendedor,
      CAST(COALESCE(h.gbv_churn_descontado_transaction,0) AS NUMERIC) AS gbv_liquido,
      CAST(COALESCE(h.atingimento_meta, 0) AS NUMERIC)                AS atingimento_meta,
      CAST(COALESCE(h.multiplicador, 0) AS NUMERIC)                   AS multiplicador,
      CAST(COALESCE(h.vlr_final_comissao, 0) AS NUMERIC)              AS vlr_final_comissao,
      CAST(COALESCE(h.total_comissao, 0) AS NUMERIC)                  AS total_comissao,
      CAST(COALESCE(mv.valor_meta, 0) AS NUMERIC)                     AS valor_meta,
      LOWER(COALESCE(mv.gestor, ''))                                  AS gestor_email,
      LOWER(COALESCE(mv.cargo, ''))                                   AS cargo
    FROM `fluency-finance.commission.vw_comissao` h
    LEFT JOIN `fluency-finance.commission.hierarquia_comercial` mv
      ON LOWER(mv.email_vendedor) = LOWER(h.vendedor)
      AND mv.mes_venda = DATE(@mes)
    WHERE h.vendedor IS NOT NULL AND h.vendedor != '#N/A'
      AND DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
    ORDER BY gbv_liquido DESC
"""


def _build_team_totals(mes: str, tl_email: str | None = None) -> dict:
    team = {"meta_total": None, "gbv_total": 0.0, "comissao_total": 0.0,
            "gbv_bruto": 0.0, "churn": 0.0,
            "coord_comissao": 0.0, "coord_email": COORDENADOR_EMAIL}
    try:
        if tl_email:
            tr = run_query("""
                SELECT
                  CAST(COALESCE(SUM(h.gbv_churn_descontado_transaction),0) AS NUMERIC) AS gbv_total,
                  CAST(COALESCE(SUM(h.gbv),0)                            AS NUMERIC) AS gbv_bruto,
                  CAST(COALESCE(SUM(h.gbv_apenas_churn_transaction),0)   AS NUMERIC) AS churn_total,
                  CAST(COALESCE(SUM(
                    CASE
                      WHEN LOWER(COALESCE(mv.cargo,'')) = 'assistente'
                        THEN COALESCE(h.atingimento_meta,0) * 4000.0 * COALESCE(h.multiplicador,0)
                      WHEN LOWER(COALESCE(mv.cargo,'')) = 'team leader'
                       AND LOWER(h.vendedor) NOT IN ('vanessa.lopes@fluencyacademy.io','tacyana.bueno@fluencyacademy.io')
                        THEN COALESCE(h.atingimento_meta,0) * 7000.0 * COALESCE(h.multiplicador,0)
                      ELSE h.vlr_final_comissao
                    END
                  ),0) AS NUMERIC) AS comissao_total
                FROM `fluency-finance.commission.vw_comissao` h
                JOIN `fluency-finance.commission.hierarquia_comercial` mv
                  ON LOWER(mv.email_vendedor) = LOWER(h.vendedor)
                  AND mv.mes_venda = DATE(@mes)
                  AND (LOWER(mv.gestor) = LOWER(@tl) OR LOWER(mv.email_vendedor) = LOWER(@tl))
                WHERE DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
            """, [bigquery.ScalarQueryParameter("mes", "DATE", mes),
                  bigquery.ScalarQueryParameter("tl",  "STRING", tl_email)], cache_ttl=90)
            mr = run_query("""
                SELECT CAST(COALESCE(SUM(valor_meta),0) AS NUMERIC) AS meta_total
                FROM `fluency-finance.commission.hierarquia_comercial`
                WHERE mes_venda = DATE(@mes) AND LOWER(gestor) = LOWER(@tl)
            """, [bigquery.ScalarQueryParameter("mes", "DATE", mes),
                  bigquery.ScalarQueryParameter("tl",  "STRING", tl_email)], cache_ttl=90)
        else:
            tr = run_query("""
                SELECT
                  CAST(COALESCE(SUM(h.gbv_churn_descontado_transaction),0) AS NUMERIC) AS gbv_total,
                  CAST(COALESCE(SUM(h.gbv),0)                            AS NUMERIC) AS gbv_bruto,
                  CAST(COALESCE(SUM(h.gbv_apenas_churn_transaction),0)   AS NUMERIC) AS churn_total,
                  CAST(COALESCE(SUM(
                    CASE
                      WHEN LOWER(COALESCE(mv.cargo,'')) = 'assistente'
                        THEN COALESCE(h.atingimento_meta,0) * 4000.0 * COALESCE(h.multiplicador,0)
                      WHEN LOWER(COALESCE(mv.cargo,'')) = 'team leader'
                       AND LOWER(h.vendedor) NOT IN ('vanessa.lopes@fluencyacademy.io','tacyana.bueno@fluencyacademy.io')
                        THEN COALESCE(h.atingimento_meta,0) * 7000.0 * COALESCE(h.multiplicador,0)
                      ELSE h.vlr_final_comissao
                    END
                  ),0) AS NUMERIC) AS comissao_total
                FROM `fluency-finance.commission.vw_comissao` h
                LEFT JOIN `fluency-finance.commission.hierarquia_comercial` mv
                  ON LOWER(mv.email_vendedor) = LOWER(h.vendedor)
                  AND mv.mes_venda = DATE(@mes)
                WHERE h.vendedor IS NOT NULL AND h.vendedor != '#N/A'
                  AND DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
            """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=90)
            mr = run_query("""
                SELECT CAST(COALESCE(SUM(valor_meta),0) AS NUMERIC) AS meta_total
                FROM `fluency-finance.commission.hierarquia_comercial`
                WHERE mes_venda = DATE(@mes)
                  AND LOWER(COALESCE(cargo,'')) NOT LIKE '%team leader%'
                  AND LOWER(COALESCE(cargo,'')) NOT LIKE '%coorden%'
            """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=90)
        if tr:
            team["gbv_total"]      = float(tr[0]["gbv_total"])
            team["gbv_bruto"]      = float(tr[0]["gbv_bruto"])
            team["churn"]          = float(tr[0]["churn_total"])
            team["comissao_total"] = float(tr[0]["comissao_total"])
        if mr:
            team["meta_total"] = float(mr[0]["meta_total"])
        # Soma extras aprovados ao GBV do time.
        try:
            if tl_email:
                er = run_query("""
                    SELECT COALESCE(SUM(CASE WHEN ev.is_churn=0 OR ev.is_churn IS NULL
                                            THEN ev.gbv ELSE 0 END), 0) AS extras_gbv
                    FROM `fluency-finance.commission.extras_vendedores` ev
                    JOIN `fluency-finance.commission.hierarquia_comercial` mv
                      ON LOWER(mv.email_vendedor) = LOWER(ev.vendedor)
                      AND mv.mes_venda = DATE(@mes)
                      AND (LOWER(mv.gestor) = LOWER(@tl) OR LOWER(mv.email_vendedor) = LOWER(@tl))
                    WHERE ev.competencia = DATE(@mes)
                      AND ev.status_tl = 'aprovado' AND ev.status_coord = 'aprovado'
                """, [bigquery.ScalarQueryParameter("mes", "DATE", mes),
                      bigquery.ScalarQueryParameter("tl",  "STRING", tl_email)], cache_ttl=90)
            else:
                er = run_query("""
                    SELECT COALESCE(SUM(CASE WHEN is_churn=0 OR is_churn IS NULL
                                            THEN gbv ELSE 0 END), 0) AS extras_gbv
                    FROM `fluency-finance.commission.extras_vendedores`
                    WHERE competencia = DATE(@mes)
                      AND status_tl = 'aprovado' AND status_coord = 'aprovado'
                """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=90)
            if er:
                extras_team = float(er[0]["extras_gbv"] or 0)
                team["gbv_total"] += extras_team
                team["gbv_bruto"] += extras_team
        except Exception:
            pass
        # comissão própria do coordenador (Fabio) — separada do total do time, p/ ele acompanhar
        cr = run_query("""
            SELECT CAST(COALESCE(SUM(vlr_final_comissao),0) AS NUMERIC) AS v
            FROM `fluency-finance.commission.vw_comissao`
            WHERE LOWER(vendedor) = LOWER(@coord)
              AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
        """, [bigquery.ScalarQueryParameter("mes",   "DATE",   mes),
              bigquery.ScalarQueryParameter("coord", "STRING", COORDENADOR_EMAIL)], cache_ttl=90)
        if cr:
            team["coord_comissao"] = float(cr[0]["v"])
    except Exception:
        pass
    return team


@app.route("/api/financial-summary")
@login_required
def api_financial_summary():
    """Resumo financeiro hierárquico (Coord → TL → membro) para o papel aprovador e master.
    Retorna dados reais do BQ com correção OTE e custo/GBV por nível."""
    role = _get_role_data()["role"]
    if role not in ("aprovador", "master", "gestor"):
        return jsonify({"error": "forbidden"}), 403
    mes = resolve_month()

    # ── 1. Todos os colaboradores do roster + comissão do mês ─────────────────
    rows = run_query("""
        SELECT
          LOWER(h.email_vendedor)                                          AS email,
          LOWER(COALESCE(h.cargo,''))                                      AS cargo,
          LOWER(COALESCE(h.gestor,''))                                     AS gestor,
          CAST(COALESCE(h.valor_meta,0)                        AS FLOAT64) AS meta,
          CAST(COALESCE(v.gbv_churn_descontado_transaction,0)  AS FLOAT64) AS gbv_liq,
          CAST(COALESCE(v.gbv,0)                               AS FLOAT64) AS gbv_bruto,
          CAST(COALESCE(v.gbv_apenas_churn_transaction,0)      AS FLOAT64) AS churn,
          CAST(COALESCE(v.atingimento_meta,0)                  AS FLOAT64) AS atingimento,
          CAST(COALESCE(v.multiplicador,0)                     AS FLOAT64) AS mult,
          CAST(COALESCE(v.vlr_final_comissao,0)                AS FLOAT64) AS vlr_final
        FROM `fluency-finance.commission.hierarquia_comercial` h
        LEFT JOIN `fluency-finance.commission.vw_comissao` v
          ON LOWER(v.vendedor) = LOWER(h.email_vendedor)
          AND DATE_TRUNC(DATE(v.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
        WHERE h.mes_venda = DATE(@mes)
          AND h.email_vendedor IS NOT NULL
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=120)

    # ── 2. Status de aceite ───────────────────────────────────────────────────
    signed_rows = run_query("""
        SELECT LOWER(vendedor) AS email
        FROM `fluency-finance.commission.signoff_vendedores`
        WHERE competencia = DATE(@mes) AND signed_at IS NOT NULL
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=30)
    signed_set = {r["email"] for r in signed_rows}

    # ── 3. Construir dicionário de membros com correção OTE ───────────────────
    members: dict[str, dict] = {}
    for r in rows:
        email = r["email"]
        cargo = r["cargo"]
        modelo = _modelo_comissao(email, cargo)
        ating  = float(r["atingimento"] or 0)
        mult   = float(r["mult"] or 0)
        vlr    = _corrige_vlr_ote(float(r["vlr_final"] or 0), modelo, ating, mult)
        newcomer = _is_newcomer_assist(email, mes)
        gbv_liq  = float(r["gbv_liq"] or 0)
        custo_gbv = (vlr / gbv_liq) if gbv_liq > 0 else 0.0
        members[email] = {
            "email":     email,
            "name":      _name_from_email(email),
            "cargo":     cargo,
            "gestor":    r["gestor"],
            "modelo":    modelo,
            "meta":      float(r["meta"] or 0),
            "gbv_liq":   round(gbv_liq, 2),
            "gbv_bruto": round(float(r["gbv_bruto"] or 0), 2),
            "churn":     round(float(r["churn"] or 0), 2),
            "atingimento": round(ating, 6),
            "mult":      round(mult, 4),
            "comissao":  round(vlr, 2),
            "custo_gbv": round(custo_gbv, 6),
            "newcomer":  newcomer,
            "signed":    (email in signed_set) and not newcomer,
        }

    # ── 4. Agregar GBV/meta por TL (base do atingimento deles) ───────────────
    tl_team_gbv:  dict[str, float] = {}
    tl_team_meta: dict[str, float] = {}
    tl_team_com:  dict[str, float] = {}
    for m in members.values():
        if "team leader" not in m["cargo"] and m["email"] != COORDENADOR_EMAIL.lower():
            tl = m["gestor"]
            tl_team_gbv[tl]  = tl_team_gbv.get(tl, 0.0)  + m["gbv_liq"]
            tl_team_meta[tl] = tl_team_meta.get(tl, 0.0) + m["meta"]
            tl_team_com[tl]  = tl_team_com.get(tl, 0.0)  + m["comissao"]

    # ── 5. Montar hierarquia ──────────────────────────────────────────────────
    coord_email = COORDENADOR_EMAIL.lower()
    coord_data  = members.get(coord_email, {
        "email": coord_email, "name": _name_from_email(COORDENADOR_EMAIL),
        "cargo": "coordenador", "modelo": "coord",
        "gbv_liq": 0.0, "comissao": 0.0, "atingimento": 0.0,
        "custo_gbv": 0.0, "newcomer": False,
        "signed": coord_email in signed_set,
    })

    tl_emails = sorted(
        [e for e, m in members.items() if "team leader" in m["cargo"]],
        key=lambda e: _name_from_email(e)
    )
    tls_out = []
    for tl_email in tl_emails:
        tl = dict(members[tl_email])
        tgbv  = tl_team_gbv.get(tl_email, 0.0)
        tmeta = tl_team_meta.get(tl_email, 0.0)
        tcom  = tl_team_com.get(tl_email, 0.0)
        tl["team_gbv"]       = round(tgbv, 2)
        tl["team_meta"]      = round(tmeta, 2)
        tl["team_comissao"]  = round(tcom, 2)
        tl["team_atingimento"] = round(tgbv / tmeta, 6) if tmeta > 0 else 0.0
        tl["team_custo_gbv"] = round(tcom / tgbv, 6) if tgbv > 0 else 0.0
        team_members = sorted(
            [m for m in members.values()
             if m["gestor"] == tl_email and "team leader" not in m["cargo"]],
            key=lambda m: -m["gbv_liq"]
        )
        tl["members"] = team_members
        tls_out.append(tl)

    # Liderados diretos do coord (sem TL como gestor)
    tl_set = set(tl_emails)
    direct = sorted(
        [m for m in members.values()
         if m["gestor"] == coord_email and "team leader" not in m["cargo"]
         and m["email"] != coord_email],
        key=lambda m: -m["gbv_liq"]
    )

    # Totais globais
    all_vend = [m for m in members.values()
                if "team leader" not in m["cargo"] and m["email"] != coord_email]
    total_gbv = sum(m["gbv_liq"] for m in all_vend)
    total_com = sum(m["comissao"] for m in members.values())
    custo_total = (total_com / total_gbv) if total_gbv > 0 else 0.0

    return jsonify({
        "coord":   coord_data,
        "tls":     tls_out,
        "direct":  direct,
        "totals": {
            "gbv_liq":   round(total_gbv, 2),
            "comissao":  round(total_com, 2),
            "custo_gbv": round(custo_total, 6),
        },
        "mes": mes,
    })


@app.route("/api/ranking")
@login_required
def api_ranking():
    email     = effective_email()
    mes       = resolve_month()
    role_data = _get_role_data()
    role      = role_data["role"]

    # Vendedores nunca veem ranking
    if role == "vendedor":
        return jsonify({"rows": [], "team": None, "visible": False})

    if role == "tl":
        rows_bq = run_query(_RANKING_SQL_TEAM, [
            bigquery.ScalarQueryParameter("mes",      "DATE",   mes),
            bigquery.ScalarQueryParameter("tl_email", "STRING", email),
        ], cache_ttl=90)
        show_full = True
        tl_filter = email

    elif role == "gestor":
        selected = request.args.get("vendedor", "").strip().lower()
        if not selected or selected == "todos":
            # Gestor overview uses /api/tl-summary; ranking hidden
            return jsonify({"rows": [], "team": _build_team_totals(mes), "visible": False})
        # Drill into a specific TL
        rows_bq   = run_query(_RANKING_SQL_TEAM, [
            bigquery.ScalarQueryParameter("mes",      "DATE",   mes),
            bigquery.ScalarQueryParameter("tl_email", "STRING", selected),
        ], cache_ttl=90)
        show_full = True
        tl_filter = selected

    else:  # master / people_ops
        rows_bq   = run_query(_RANKING_SQL_ALL, [
            bigquery.ScalarQueryParameter("mes", "DATE", mes),
        ], cache_ttl=90)
        show_full = True
        tl_filter = None

    rows_bq = _filter_active(rows_bq, "vendedor")

    # Extras aprovados do mês (uma query para todos os vendedores do ranking).
    extras_bulk = _approved_extras_bulk(mes)

    result = []
    for r in rows_bq:
        vend_email = r["vendedor"]
        is_me      = vend_email.lower() == email.lower()
        label      = _name_from_email(vend_email) if (show_full or is_me) \
                     else _initial_from_email(vend_email)
        gestor_raw = str(r.get("gestor_email", "") or "")

        gbv_liquido    = float(r["gbv_liquido"])
        atingimento    = float(r["atingimento_meta"])
        multiplicador  = float(r["multiplicador"])
        cargo_r        = str(r.get("cargo") or "")
        comissao_final = _corrige_vlr_ote(
            float(r["vlr_final_comissao"]), _modelo_comissao(vend_email, cargo_r),
            atingimento, multiplicador)
        valor_meta     = float(r.get("valor_meta") or 0)

        # Aplica extras aprovados: recalcula GBV, atingimento, mult e comissão.
        vextras = [e for e in extras_bulk.get(vend_email.lower(), []) if not e["is_churn"]]
        if vextras:
            extras_gbv = sum(e["gbv"] for e in vextras)
            gbv_liquido += extras_gbv
            if valor_meta > 0:
                atingimento = gbv_liquido / valor_meta
            modelo  = _modelo_comissao(vend_email, cargo_r)
            new_mult = _mult_for_modelo(modelo, atingimento)
            if new_mult is not None:
                multiplicador = new_mult
                if modelo in _RATES_BY_MODELO:
                    rates     = _RATES_BY_MODELO[modelo]
                    extra_com = sum(e["gbv"] * rates.get(e["modality_payment"], 0) for e in vextras)
                    comissao_final = (float(r.get("total_comissao") or 0) + extra_com) * new_mult
                elif modelo in _OTE_BY_MODELO:
                    comissao_final = atingimento * _OTE_BY_MODELO[modelo] * new_mult

        result.append({
            "posicao":        int(r["posicao"]),
            "isMe":           is_me,
            "label":          label,
            "vendedor":       vend_email if role in READ_ALL_ROLES else None,
            "gbv_liquido":    gbv_liquido,
            "atingimento":    atingimento,
            "multiplicador":  multiplicador,
            "comissao_final": comissao_final,
            "valor_meta":     valor_meta if role in READ_ALL_ROLES else 0,
            "gestor_label":   _name_from_email(gestor_raw) if (role in READ_ALL_ROLES and gestor_raw and not gestor_raw.startswith("#")) else None,
            "nota_fim_mes":   _is_newcomer_assist(vend_email, mes),
        })

    team = _build_team_totals(mes, tl_filter)
    return jsonify({"rows": result, "team": team, "visible": True})


# ── API: TL summary (gestor view) ─────────────────────────────────────────────

@app.route("/api/tl-summary")
@login_required
def api_tl_summary():
    role_data = _get_role_data()
    if role_data["role"] not in ("gestor", "master", "people_ops"):
        return jsonify({"error": "forbidden"}), 403
    mes = resolve_month()
    sql = """
        WITH tl_teams AS (
          SELECT
            LOWER(mv_tl.email_vendedor) AS tl_email,
            LOWER(mv_m.email_vendedor)  AS member_email,
            mv_m.valor_meta             AS member_meta
          FROM `fluency-finance.commission.hierarquia_comercial` mv_tl
          JOIN `fluency-finance.commission.hierarquia_comercial` mv_m
            ON LOWER(mv_m.gestor) = LOWER(mv_tl.email_vendedor)
            AND mv_m.mes_venda = DATE(@mes)
            AND LOWER(COALESCE(mv_m.cargo,'')) NOT LIKE '%team leader%'
          WHERE mv_tl.mes_venda = DATE(@mes)
            AND LOWER(COALESCE(mv_tl.cargo,'')) LIKE '%team leader%'
        ),
        tl_own AS (
          SELECT
            LOWER(h.vendedor)                                         AS tl_email,
            CAST(COALESCE(h.gbv_churn_descontado_transaction,0) AS NUMERIC) AS tl_gbv,
            CAST(COALESCE(h.atingimento_meta,0) AS NUMERIC)           AS tl_ating,
            CAST(COALESCE(h.multiplicador,0) AS NUMERIC)              AS tl_mult,
            CAST(COALESCE(h.vlr_final_comissao,0) AS NUMERIC)         AS tl_comissao
          FROM `fluency-finance.commission.vw_comissao` h
          WHERE DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
            AND h.vendedor IS NOT NULL AND h.vendedor != '#N/A'
        )
        SELECT
          t.tl_email,
          CAST(COALESCE(SUM(h.gbv_churn_descontado_transaction),0) AS NUMERIC) AS team_gbv,
          CAST(COALESCE(SUM(h.vlr_final_comissao),0) AS NUMERIC)               AS team_comissao,
          CAST(COALESCE(SUM(t.member_meta),0) AS NUMERIC)                      AS team_meta,
          COUNT(DISTINCT t.member_email)                                        AS member_count,
          MAX(o.tl_gbv)      AS tl_own_gbv,
          MAX(o.tl_ating)    AS tl_own_ating,
          MAX(o.tl_mult)     AS tl_own_mult,
          MAX(o.tl_comissao) AS tl_own_comissao
        FROM tl_teams t
        LEFT JOIN `fluency-finance.commission.vw_comissao` h
          ON LOWER(h.vendedor) = t.member_email
          AND DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
        LEFT JOIN tl_own o ON o.tl_email = t.tl_email
        GROUP BY t.tl_email
        ORDER BY team_gbv DESC
    """
    rows = run_query(sql, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=90)
    rows = _filter_active(rows, "tl_email")

    # Extras aprovados do mês — agrupados por membro para somar ao GBV de cada time.
    extras_bulk = _approved_extras_bulk(mes)
    # Mapeia gestor → membros usando o roster (para saber quem pertence a qual TL).
    vmap = _load_vmap()
    tl_members: dict[str, list[str]] = {}
    for em, info in vmap.items():
        gestor = (info.get("gestor") or "").lower()
        if gestor:
            tl_members.setdefault(gestor, []).append(em.lower())

    result = []
    for r in rows:
        tl = str(r["tl_email"])
        team_gbv = float(r["team_gbv"])
        # Soma o GBV dos extras aprovados de cada membro deste time.
        members_gbv_extra = sum(
            e["gbv"]
            for m in tl_members.get(tl, [])
            for e in extras_bulk.get(m, [])
            if not e["is_churn"]
        )
        result.append({
            "tl_email":        tl,
            "label":           _name_from_email(tl),
            "team_gbv":        team_gbv + members_gbv_extra,
            "team_comissao":   float(r["team_comissao"]),
            "team_meta":       float(r["team_meta"]),
            "member_count":    int(r["member_count"]),
            "tl_own_gbv":      float(r["tl_own_gbv"])      if r["tl_own_gbv"]      is not None else 0.0,
            "tl_own_ating":    float(r["tl_own_ating"])    if r["tl_own_ating"]    is not None else 0.0,
            "tl_own_mult":     float(r["tl_own_mult"])     if r["tl_own_mult"]     is not None else 0.0,
            "tl_own_comissao": float(r["tl_own_comissao"]) if r["tl_own_comissao"] is not None else 0.0,
        })
    return jsonify(result)


# ── API: Produtividade por horário (master + gestor) ──────────────────────────

_PRODUTIVIDADE_HORA_SQL = """
    WITH vendas AS (
      SELECT
        LOWER(vendedor)             AS vendedor,
        is_churn,
        gbv,
        contract_created_at_brt
      FROM `fluency-gold.conversion.obt_conversions`
      WHERE DATE_TRUNC(contract_created_at_brt_date, MONTH) = DATE(@mes)
        AND vendedor IS NOT NULL
        AND (@dia IS NULL OR EXTRACT(DAY FROM contract_created_at_brt_date) = @dia)
    ),
    hier AS (
      -- Squad = o próprio TL quando a pessoa é TL (venda própria conta no squad dela);
      -- senão, squad = o gestor (TL) dela. Sem mapa fixo de nome de squad — o roster manda.
      SELECT
        LOWER(email_vendedor) AS email_vendedor,
        CASE WHEN LOWER(COALESCE(cargo,'')) LIKE '%team leader%'
             THEN LOWER(email_vendedor)
             ELSE LOWER(COALESCE(gestor,'')) END AS squad_lead
      FROM `fluency-finance.commission.hierarquia_comercial`
      WHERE mes_venda = DATE(@mes)
    )
    SELECT
      COALESCE(NULLIF(h.squad_lead,''), 'sem_time') AS squad_lead,
      v.vendedor,
      EXTRACT(HOUR FROM v.contract_created_at_brt)  AS hora,
      COUNT(*)                                       AS qtd_vendas,
      SUM(IF(v.is_churn, 0, v.gbv))                  AS gbv_pc
    FROM vendas v
    LEFT JOIN hier h ON h.email_vendedor = v.vendedor
    GROUP BY 1,2,3
"""

def _empty_horas() -> list[dict]:
    return [{"hora": h, "gbv": 0.0, "qtd": 0, "ticket_medio": 0.0} for h in range(24)]

@app.route("/api/produtividade-hora")
@login_required
def api_produtividade_hora():
    role = _get_role_data()["role"]
    if role not in ("master", "gestor"):
        return jsonify({"error": "forbidden"}), 403
    mes = resolve_month()
    dia_raw = request.args.get("dia", "").strip()
    dia = int(dia_raw) if dia_raw.isdigit() and 1 <= int(dia_raw) <= 31 else None
    rows = run_query(
        _PRODUTIVIDADE_HORA_SQL,
        [
            bigquery.ScalarQueryParameter("mes", "DATE", mes),
            bigquery.ScalarQueryParameter("dia", "INT64", dia),
        ],
        cache_ttl=120,
    )

    times: dict[str, dict] = {}
    vendedores: dict[str, dict] = {}
    for r in rows:
        squad = str(r["squad_lead"])
        vend  = str(r["vendedor"])
        hora  = int(r["hora"])
        qtd   = int(r["qtd_vendas"])
        gbv   = float(r["gbv_pc"] or 0)

        t = times.setdefault(squad, {
            "tl_email": squad,
            "label": "Sem time atribuído" if squad == "sem_time" else _name_from_email(squad),
            "horas": _empty_horas(),
        })
        t["horas"][hora]["qtd"] += qtd
        t["horas"][hora]["gbv"] += gbv

        v = vendedores.setdefault(vend, {
            "squad_lead": squad,
            "label": _name_from_email(vend),
            "horas": _empty_horas(),
        })
        v["horas"][hora]["qtd"] += qtd
        v["horas"][hora]["gbv"] += gbv

    for bucket in (*times.values(), *vendedores.values()):
        for h in bucket["horas"]:
            h["ticket_medio"] = (h["gbv"] / h["qtd"]) if h["qtd"] else 0.0

    times_out = _filter_active(list(times.values()), "tl_email")
    # _filter_active espera uma lista de dicts com a chave de e-mail; o dict de
    # vendedores é indexado por e-mail, então filtramos as chaves inativas direto.
    inactive = _inactive_emails_for_month(mes) or set()
    vendedores_out = {
        email: v for email, v in vendedores.items()
        if email.lower() not in inactive
    }

    return jsonify({
        "mes": mes,
        "dia": dia,
        "times": sorted(times_out, key=lambda t: t["label"]),
        "vendedores": vendedores_out,
    })


def _build_leadership_payload(mes: str) -> dict:
    """Retorna {coord, tls} com status de sign-off de coord e TLs do mês.
    authorized=True sempre (signoff_autorizacao não existe — tudo liberado)."""
    rows = run_query("""
        SELECT DISTINCT
          LOWER(h.gestor) AS email,
          COALESCE(e.full_name, '') AS full_name,
          (SELECT STRING_AGG(DISTINCT h2.time, ' / ')
             FROM `fluency-finance.commission.hierarquia_comercial` h2
            WHERE LOWER(h2.gestor) = LOWER(h.gestor) AND h2.mes_venda = @mes
              AND LOWER(COALESCE(h2.cargo,'')) NOT LIKE '%team leader%') AS team
        FROM `fluency-finance.commission.hierarquia_comercial` h
        LEFT JOIN `fluency-finance.commission.employees_view` e
          ON LOWER(e.email) = LOWER(h.gestor)
        WHERE h.mes_venda = @mes
          AND h.gestor IS NOT NULL
          AND LOWER(h.gestor) != LOWER(@coord)
          AND LOWER(COALESCE(h.cargo,'')) IN ('vendedor', 'assistente')
    """, [
        bigquery.ScalarQueryParameter("mes",   "DATE",   mes),
        bigquery.ScalarQueryParameter("coord", "STRING", COORDENADOR_EMAIL),
    ], cache_ttl=300)

    tl_emails = [r["email"] for r in rows if r.get("email")]
    all_emails = [COORDENADOR_EMAIL.lower()] + tl_emails

    signed_rows = run_query("""
        SELECT LOWER(vendedor) AS email, signed_at
        FROM `fluency-finance.commission.signoff_vendedores`
        WHERE LOWER(vendedor) IN UNNEST(@emails) AND competencia = @mes
    """, [
        bigquery.ArrayQueryParameter("emails", "STRING", all_emails),
        bigquery.ScalarQueryParameter("mes",   "DATE",   mes),
    ], cache_ttl=30)
    signed_map = {r["email"]: r["signed_at"] for r in signed_rows}

    def _hydrate(email, name, team=None):
        sat = signed_map.get(email.lower())
        return {
            "email": email,
            "name": name or _name_from_email(email),
            "team": team,
            "signed": sat is not None,
            "signed_at": str(sat) if sat else None,
            "authorized": True,
        }

    coord_entry = _hydrate(COORDENADOR_EMAIL, "Fabio Dias")
    tls_list = [
        _hydrate(r["email"], r.get("full_name") or "", r.get("team"))
        for r in rows if r.get("email")
    ]
    return {"coord": coord_entry, "tls": tls_list}

@app.route("/api/signoff-status")
@login_required
def api_signoff_status():
    """Status de aceite por liderado, agrupado por TL.
    TL: só o próprio time. Gestor/master/people_ops: todos os times."""
    role_data = _get_role_data()
    role = role_data["role"]
    me   = effective_email()
    mes  = resolve_month()

    if role not in ("gestor", "master", "people_ops", "aprovador", "tl"):
        return jsonify({"error": "forbidden"}), 403

    params = [bigquery.ScalarQueryParameter("mes", "DATE", mes)]
    tl_filter = ""
    if role == "tl":
        tl_filter = "AND LOWER(h.gestor) = LOWER(@tl)"
        params.append(bigquery.ScalarQueryParameter("tl", "STRING", me))

    rows = run_query(f"""
        SELECT
          LOWER(h.gestor)         AS tl_email,
          LOWER(h.email_vendedor) AS member_email,
          s.signed_at
        FROM `fluency-finance.commission.hierarquia_comercial` h
        LEFT JOIN `fluency-finance.commission.signoff_vendedores` s
          ON LOWER(s.vendedor) = LOWER(h.email_vendedor)
          AND s.competencia = DATE(@mes)
        WHERE h.mes_venda = DATE(@mes)
          AND LOWER(COALESCE(h.cargo,'')) NOT LIKE '%team leader%'
          AND LOWER(COALESCE(h.cargo,'')) NOT LIKE '%coorden%'
          {tl_filter}
        ORDER BY h.gestor, s.signed_at IS NULL, h.email_vendedor
    """, params, cache_ttl=30)

    by_tl: dict = {}
    for r in rows:
        tl = str(r["tl_email"])
        if tl not in by_tl:
            by_tl[tl] = {"tl_email": tl, "label": _name_from_email(tl), "signed": [], "unsigned": []}
        member = {"email": str(r["member_email"]), "label": _name_from_email(str(r["member_email"]))}
        if r.get("signed_at") and not _is_newcomer_assist(str(r["member_email"]), mes):
            member["signed_at"] = str(r["signed_at"])
            by_tl[tl]["signed"].append(member)
        else:
            by_tl[tl]["unsigned"].append(member)

    result = list(by_tl.values())
    total        = sum(len(t["signed"]) + len(t["unsigned"]) for t in result)
    total_signed = sum(len(t["signed"]) for t in result)
    leadership = {}
    if role in ("gestor", "master", "aprovador"):
        try:
            leadership = _build_leadership_payload(mes)
            # Adiciona o aprovador final como entrada separada no bloco de liderança.
            aprov_signed = run_query("""
                SELECT signed_at FROM `fluency-finance.commission.signoff_vendedores`
                WHERE LOWER(vendedor) = @aprov AND competencia = DATE(@mes) LIMIT 1
            """, [bigquery.ScalarQueryParameter("aprov", "STRING", APROVADOR_FINAL_EMAIL.lower()),
                  bigquery.ScalarQueryParameter("mes",   "DATE",   mes)], cache_ttl=30)
            sat = aprov_signed[0]["signed_at"] if aprov_signed else None
            leadership["aprovador"] = {
                "email": APROVADOR_FINAL_EMAIL.lower(),
                "name": _name_from_email(APROVADOR_FINAL_EMAIL),
                "signed": sat is not None,
                "signed_at": str(sat) if sat else None,
            }
        except Exception as e:
            app.logger.warning("_build_leadership_payload falhou: %s", e)
    elif role == "tl":
        # TL vê a si mesmo no bloco de liderança (sem coord, só o próprio entry)
        try:
            signed_rows = run_query("""
                SELECT signed_at FROM `fluency-finance.commission.signoff_vendedores`
                WHERE LOWER(vendedor) = LOWER(@v) AND competencia = DATE(@mes) LIMIT 1
            """, [bigquery.ScalarQueryParameter("v", "STRING", me),
                  bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=30)
            sat = signed_rows[0]["signed_at"] if signed_rows else None
            team = None
            tl_entry = {
                "email": me.lower(),
                "name": _name_from_email(me),
                "team": team,
                "signed": sat is not None,
                "signed_at": str(sat) if sat else None,
                "authorized": True,
            }
            leadership = {"coord": None, "tls": [tl_entry]}
        except Exception as e:
            app.logger.warning("_build_tl_leadership falhou: %s", e)
    return jsonify({"by_tl": result, "total": total, "total_signed": total_signed,
                    "leadership": leadership})

# ── DSR helpers ───────────────────────────────────────────────────────────────

_MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]


def _dsr_factor(year: int, month: int) -> float:
    """Returns DSR coefficient = sundays / (total_days - sundays) for payment month."""
    _, total = calendar.monthrange(year, month)
    sundays  = sum(1 for d in range(1, total + 1)
                   if date(year, month, d).weekday() == 6)
    working  = total - sundays
    return sundays / working if working else 0.0


def _payment_month(competencia: str) -> tuple[int, int]:
    """Returns (year, month) of payment for a given competence date string YYYY-MM-01."""
    d = date.fromisoformat(competencia)
    if d.month == 12:
        return d.year + 1, 1
    return d.year, d.month + 1


# ── Preview: TL view (master only, server-rendered) ──────────────────────────

@app.route("/preview")
@login_required
def tl_preview():
    if not is_master():
        return "Acesso restrito — master only", 403

    email = request.args.get("email", "tacyana.bueno@fluencyacademy.io").strip().lower()

    # Resolve month
    mes = request.args.get("mes", "").strip()
    if not mes:
        rows = run_query("""
            SELECT MAX(DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH)) AS mes
            FROM `fluency-finance.commission.vw_comissao`
            WHERE vendedor IS NOT NULL AND vendedor != '#N/A'
        """)
        mes = str(rows[0]["mes"]) if rows and rows[0]["mes"] else date.today().replace(day=1).isoformat()

    # Individual summary
    sum_rows = run_query("""
        SELECT
          vendedor,
          CAST(COALESCE(gbv_churn_descontado_transaction, 0) AS NUMERIC) AS gbv_liquido,
          CAST(COALESCE(atingimento_meta, 0) AS NUMERIC)                 AS atingimento_meta,
          CAST(COALESCE(multiplicador, 0) AS NUMERIC)                    AS multiplicador,
          CAST(COALESCE(vlr_final_comissao, 0) AS NUMERIC)              AS vlr_final_comissao,
          CAST(COALESCE(total_comissao, 0) AS NUMERIC)                   AS total_comissao
        FROM `fluency-finance.commission.vw_comissao`
        WHERE LOWER(vendedor) = LOWER(@email)
          AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
          AND vendedor IS NOT NULL AND vendedor != '#N/A'
        LIMIT 1
    """, [bigquery.ScalarQueryParameter("email", "STRING", email),
          bigquery.ScalarQueryParameter("mes",   "DATE",   mes)])

    summary = {}
    if sum_rows:
        for k, v in sum_rows[0].items():
            summary[k] = float(v) if hasattr(v, "__float__") else v

    # Meta from meta_vendedores
    meta_rows = run_query("""
        SELECT CAST(COALESCE(valor_meta, 0) AS NUMERIC) AS valor_meta
        FROM `fluency-finance.commission.hierarquia_comercial`
        WHERE LOWER(email_vendedor) = LOWER(@email) AND mes_venda = DATE(@mes)
        LIMIT 1
    """, [bigquery.ScalarQueryParameter("email", "STRING", email),
          bigquery.ScalarQueryParameter("mes",   "DATE",   mes)])
    valor_meta = float(meta_rows[0]["valor_meta"]) if meta_rows else 0.0

    # Team ranking
    rank_rows = run_query(_RANKING_SQL_TEAM, [
        bigquery.ScalarQueryParameter("mes",      "DATE",   mes),
        bigquery.ScalarQueryParameter("tl_email", "STRING", email),
    ])
    ranking = []
    for r in rank_rows:
        ranking.append({
            "posicao":         int(r["posicao"]),
            "vendedor":        r["vendedor"],
            "name":            _name_from_email(r["vendedor"]),
            "is_tl":           r["vendedor"].lower() == email,
            "gbv_liquido":     float(r["gbv_liquido"]),
            "atingimento_meta": float(r["atingimento_meta"]),
            "multiplicador":   float(r["multiplicador"]),
            "vlr_final_comissao": float(r["vlr_final_comissao"]),
        })

    # Month label
    mes_date = date.fromisoformat(mes)
    mes_label = mes_date.strftime("%B %Y").capitalize()

    return render_template("tl_preview.html",
                           name=_name_from_email(email),
                           email=email,
                           mes=mes,
                           mes_label=mes_label,
                           summary=summary,
                           valor_meta=valor_meta,
                           ranking=ranking)


# ── API: trend (last 6 months) ────────────────────────────────────────────────

@app.route("/api/trend")
@login_required
def api_trend():
    if _get_role_data()["role"] not in ("master", "gestor", "people_ops"):
        return jsonify({"error": "forbidden"}), 403

    team_rows = run_query("""
        SELECT
          DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) AS mes,
          CAST(COALESCE(SUM(gbv_churn_descontado_transaction),0) AS NUMERIC) AS gbv_total,
          CAST(COALESCE(SUM(vlr_final_comissao),0) AS NUMERIC)               AS comissao_total,
          CAST(COALESCE(SUM(total_comissao),0) AS NUMERIC)                   AS comissao_bruta
        FROM `fluency-finance.commission.vw_comissao`
        WHERE vendedor IS NOT NULL AND vendedor != '#N/A'
          AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH)
              >= DATE_TRUNC(DATE_SUB(CURRENT_DATE('America/Sao_Paulo'), INTERVAL 5 MONTH), MONTH)
        GROUP BY mes
        ORDER BY mes ASC
    """, cache_ttl=90)

    tl_rows = run_query("""
        WITH latest_map AS (
          SELECT DISTINCT
            LOWER(mv_m.email_vendedor) AS member_email,
            LOWER(mv_tl.email_vendedor) AS tl_email
          FROM `fluency-finance.commission.hierarquia_comercial` mv_tl
          JOIN `fluency-finance.commission.hierarquia_comercial` mv_m
            ON LOWER(mv_m.gestor) = LOWER(mv_tl.email_vendedor)
            AND mv_m.mes_venda = mv_tl.mes_venda
            AND LOWER(COALESCE(mv_m.cargo,'')) NOT LIKE '%team leader%'
          WHERE LOWER(COALESCE(mv_tl.cargo,'')) LIKE '%team leader%'
            AND mv_tl.mes_venda = (SELECT MAX(mes_venda)
                                   FROM `fluency-finance.commission.hierarquia_comercial`)
        )
        SELECT
          DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) AS mes,
          m.tl_email,
          CAST(COALESCE(SUM(h.gbv_churn_descontado_transaction),0) AS NUMERIC) AS gbv_total,
          CAST(COALESCE(SUM(h.vlr_final_comissao),0) AS NUMERIC)               AS comissao_total
        FROM `fluency-finance.commission.vw_comissao` h
        JOIN latest_map m ON LOWER(h.vendedor) = m.member_email
        WHERE DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH)
              >= DATE_TRUNC(DATE_SUB(CURRENT_DATE('America/Sao_Paulo'), INTERVAL 5 MONTH), MONTH)
          AND h.vendedor IS NOT NULL AND h.vendedor != '#N/A'
        GROUP BY mes, m.tl_email
        ORDER BY mes ASC, m.tl_email
    """, cache_ttl=90)

    months = sorted({str(r["mes"]) for r in team_rows})
    team_map = {str(r["mes"]): {
        "gbv":             float(r["gbv_total"]),
        "comissao":        float(r["comissao_total"]),
        "comissao_bruta":  float(r["comissao_bruta"]),
    } for r in team_rows}

    by_tl: dict[str, dict] = {}
    for r in tl_rows:
        tl = str(r["tl_email"])
        if tl not in by_tl:
            by_tl[tl] = {"label": _name_from_email(tl), "data": {}}
        by_tl[tl]["data"][str(r["mes"])] = {
            "gbv":      float(r["gbv_total"]),
            "comissao": float(r["comissao_total"]),
        }

    return jsonify({
        "months": months,
        "team":   [{"mes": m, **team_map.get(m, {"gbv": 0, "comissao": 0, "comissao_bruta": 0})}
                   for m in months],
        "by_tl":  {tl: {
                       "label":    info["label"],
                       "gbv":      [info["data"].get(m, {}).get("gbv", 0)      for m in months],
                       "comissao": [info["data"].get(m, {}).get("comissao", 0) for m in months],
                   } for tl, info in by_tl.items()},
    })


# ── API: month-over-month comparison ─────────────────────────────────────────

@app.route("/api/mom-compare")
@login_required
def api_mom_compare():
    if _get_role_data()["role"] not in ("master", "gestor", "people_ops"):
        return jsonify({"error": "forbidden"}), 403

    team_rows = run_query("""
        SELECT
          DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) AS mes,
          CAST(COALESCE(SUM(gbv_churn_descontado_transaction),0) AS NUMERIC) AS gbv,
          CAST(COALESCE(SUM(total_comissao),0) AS NUMERIC)                   AS com_bruta,
          CAST(COALESCE(SUM(vlr_final_comissao),0) AS NUMERIC)               AS com_final
        FROM `fluency-finance.commission.vw_comissao`
        WHERE vendedor IS NOT NULL AND vendedor != '#N/A'
          AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH)
              >= DATE_TRUNC(DATE_SUB(CURRENT_DATE('America/Sao_Paulo'), INTERVAL 5 MONTH), MONTH)
        GROUP BY mes
        ORDER BY mes ASC
    """, cache_ttl=90)

    tl_rows = run_query("""
        WITH latest_map AS (
          SELECT DISTINCT
            LOWER(mv_m.email_vendedor) AS member_email,
            LOWER(mv_tl.email_vendedor) AS tl_email
          FROM `fluency-finance.commission.hierarquia_comercial` mv_tl
          JOIN `fluency-finance.commission.hierarquia_comercial` mv_m
            ON LOWER(mv_m.gestor) = LOWER(mv_tl.email_vendedor)
            AND mv_m.mes_venda = mv_tl.mes_venda
            AND LOWER(COALESCE(mv_m.cargo,'')) NOT LIKE '%team leader%'
          WHERE LOWER(COALESCE(mv_tl.cargo,'')) LIKE '%team leader%'
            AND mv_tl.mes_venda = (SELECT MAX(mes_venda)
                                   FROM `fluency-finance.commission.hierarquia_comercial`)
        )
        SELECT
          DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) AS mes,
          m.tl_email,
          CAST(COALESCE(SUM(h.gbv_churn_descontado_transaction),0) AS NUMERIC) AS gbv,
          CAST(COALESCE(SUM(h.vlr_final_comissao),0) AS NUMERIC)               AS com_final
        FROM `fluency-finance.commission.vw_comissao` h
        JOIN latest_map m ON LOWER(h.vendedor) = m.member_email
        WHERE DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH)
              >= DATE_TRUNC(DATE_SUB(CURRENT_DATE('America/Sao_Paulo'), INTERVAL 5 MONTH), MONTH)
          AND h.vendedor IS NOT NULL AND h.vendedor != '#N/A'
        GROUP BY mes, m.tl_email
        ORDER BY mes ASC, m.tl_email
    """, cache_ttl=90)

    months = sorted({str(r["mes"]) for r in team_rows})
    team = [{"mes": str(r["mes"]), "gbv": float(r["gbv"]),
             "com_bruta": float(r["com_bruta"]), "com_final": float(r["com_final"])}
            for r in team_rows]

    by_tl_map: dict[str, dict] = {}
    for r in tl_rows:
        tl = str(r["tl_email"])
        if tl not in by_tl_map:
            by_tl_map[tl] = {"label": _name_from_email(tl), "data": {}}
        by_tl_map[tl]["data"][str(r["mes"])] = float(r["com_final"])

    by_tl = [{"tl": tl, "label": info["label"],
               "com_final": [info["data"].get(m, 0) for m in months]}
              for tl, info in by_tl_map.items()]

    return jsonify({"months": months, "team": team, "by_tl": by_tl})


# ── API: payroll impact ───────────────────────────────────────────────────────

def _serialize_access(rows):
    out = []
    for r in rows:
        d = dict(r)
        if d.get("ultimo_acesso") is not None: d["ultimo_acesso"] = str(d["ultimo_acesso"])
        if d.get("accessed_at")  is not None: d["accessed_at"]  = str(d["accessed_at"])
        if d.get("acessos")      is not None: d["acessos"]      = int(d["acessos"])
        out.append(d)
    return out

@app.route("/api/heartbeat", methods=["POST"])
@login_required
def api_heartbeat():
    """Ping de presença — grava uma marca leve no access_log (path=/api/heartbeat)."""
    _log_access(_real_role())
    return jsonify({"ok": True})

@app.route("/api/presence")
@login_required
def api_presence():
    """Quem está online (acesso nos últimos 5 min). Só papéis de gestão veem."""
    if _real_role() not in ("master", "people_ops", "gestor"):
        return jsonify([])
    rows = run_query("""
        WITH recente AS (
          SELECT email, accessed_at, name, picture, role,
                 ROW_NUMBER() OVER (PARTITION BY email ORDER BY accessed_at DESC) AS rn
          FROM `fluency-finance.commission.access_log`
          WHERE accessed_at >= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 5 MINUTE)
            AND email IS NOT NULL AND email != ''
        )
        SELECT email, name, picture, role, accessed_at AS ultimo_acesso
        FROM recente WHERE rn = 1
        ORDER BY ultimo_acesso DESC
    """, cache_ttl=20)
    return jsonify(_serialize_access(rows))

@app.route("/api/access-log")
@login_required
def api_access_log():
    if not is_master():
        return jsonify({"error": "forbidden"}), 403
    by_user = run_query("""
        SELECT
          email,
          MAX(IF(path='/', accessed_at, NULL))                         AS ultimo_acesso,
          COUNTIF(path='/')                                            AS acessos,
          ARRAY_AGG(role    ORDER BY accessed_at DESC LIMIT 1)[OFFSET(0)] AS role,
          ARRAY_AGG(name    IGNORE NULLS ORDER BY accessed_at DESC LIMIT 1) AS name_arr,
          ARRAY_AGG(picture IGNORE NULLS ORDER BY accessed_at DESC LIMIT 1) AS pic_arr
        FROM `fluency-finance.commission.access_log`
        WHERE email IS NOT NULL AND email != ''
        GROUP BY email
        HAVING acessos > 0
        ORDER BY ultimo_acesso DESC
    """)
    for r in by_user:
        r["name"]    = (r.pop("name_arr", None) or [None])[0]
        r["picture"] = (r.pop("pic_arr", None) or [None])[0]
    recent = run_query("""
        SELECT email, accessed_at, role, ip, name, picture
        FROM `fluency-finance.commission.access_log`
        WHERE path = '/'
        ORDER BY accessed_at DESC
        LIMIT 100
    """)
    return jsonify({"by_user": _serialize_access(by_user), "recent": _serialize_access(recent)})

@app.route("/api/payroll-impact")
@login_required
def api_payroll_impact():
    if _get_role_data()["role"] != "master":
        return jsonify({"error": "forbidden"}), 403

    mes = resolve_month()   # competence month
    pay_year, pay_month = _payment_month(mes)

    # DSR + reflexo: parametros de folha por mês (planilha oficial "Comissao" do FP&A,
    # carregada em parametros_folha). DSR varia por mês; reflexo (encargos) = 50% sobre
    # (comissão + DSR). Fallback p/ DSR calendário se o mês não estiver na tabela.
    par = run_query("""
        SELECT dsr_rate, reflexo_rate
        FROM `fluency-finance.commission.parametros_folha`
        WHERE mes_competencia = DATE(@mes) LIMIT 1
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=90)
    if par:
        dsr_fact     = float(par[0]["dsr_rate"])
        reflexo_rate = float(par[0]["reflexo_rate"])
    else:
        dsr_fact     = _dsr_factor(pay_year, pay_month)
        reflexo_rate = 0.50

    _, total_days = calendar.monthrange(pay_year, pay_month)
    sundays  = sum(1 for d_ in range(1, total_days + 1)
                   if date(pay_year, pay_month, d_).weekday() == 6)
    working  = total_days - sundays

    rows = run_query("""
        SELECT
          h.vendedor,
          CAST(COALESCE(h.vlr_final_comissao, 0) AS NUMERIC) AS comissao_final,
          CAST(COALESCE(h.total_comissao,      0) AS NUMERIC) AS total_comissao,
          CAST(COALESCE(h.atingimento_meta,    0) AS FLOAT64) AS atingimento_meta,
          CAST(COALESCE(mv.valor_meta,         0) AS FLOAT64) AS valor_meta,
          LOWER(COALESCE(mv.cargo,            '')) AS cargo
        FROM `fluency-finance.commission.vw_comissao` h
        LEFT JOIN `fluency-finance.commission.hierarquia_comercial` mv
          ON LOWER(mv.email_vendedor) = LOWER(h.vendedor)
          AND mv.mes_venda = DATE(@mes)
        WHERE h.vendedor IS NOT NULL AND h.vendedor != '#N/A'
          AND DATE_TRUNC(DATE(h.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
        ORDER BY comissao_final DESC
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=90)

    # Extras aprovados — mesmo recálculo do ranking/summary.
    extras_bulk = _approved_extras_bulk(mes)

    # NÃO filtra ativos: folha inclui TODOS que recebem comissão, inclusive Fabio (PJ).
    vendedores = []
    tot_com = tot_dsr = tot_reflexo = 0.0
    for r in rows:
        com = float(r["comissao_final"])

        vextras = [e for e in extras_bulk.get(str(r["vendedor"]).lower(), []) if not e["is_churn"]]
        if vextras:
            vm         = float(r["valor_meta"])
            ating_base = float(r["atingimento_meta"])
            gbv_orig   = ating_base * vm if vm > 0 else 0.0
            new_ating  = (gbv_orig + sum(e["gbv"] for e in vextras)) / vm if vm > 0 else ating_base
            modelo     = _modelo_comissao(str(r["vendedor"]), str(r["cargo"]))
            new_mult   = _mult_for_modelo(modelo, new_ating)
            if new_mult is not None:
                if modelo in _RATES_BY_MODELO:
                    rates     = _RATES_BY_MODELO[modelo]
                    extra_com = sum(e["gbv"] * rates.get(e["modality_payment"], 0) for e in vextras)
                    com       = (float(r["total_comissao"]) + extra_com) * new_mult
                elif modelo in _OTE_BY_MODELO:
                    com = new_ating * _OTE_BY_MODELO[modelo] * new_mult

        if com <= 0:
            continue
        dsr     = com * dsr_fact
        reflexo = (com + dsr) * reflexo_rate
        vendedores.append({
            "label":    _name_from_email(str(r["vendedor"])),
            "comissao": com,
            "dsr":      dsr,
            "reflexo":  reflexo,
            "total":    com + dsr + reflexo,
        })
        tot_com     += com
        tot_dsr     += dsr
        tot_reflexo += reflexo

    mes_d = date.fromisoformat(mes)
    return jsonify({
        "mes_competencia_label": _MESES_PT[mes_d.month - 1] + " " + str(mes_d.year),
        "mes_pagamento_label":   _MESES_PT[pay_month - 1]   + " " + str(pay_year),
        "dsr_factor":            dsr_fact,
        "payment_sundays":       sundays,
        "payment_working_days":  working,
        "vendedores":            vendedores,
        "reflexo_rate":          reflexo_rate,
        "totais": {
            "comissao": tot_com,
            "dsr":      tot_dsr,
            "reflexo":  tot_reflexo,
            "total":    tot_com + tot_dsr + tot_reflexo,
        },
    })


# ── API: Assistente × Vendedor (rentabilidade — master only) ───────────────────
def _av_baseline():
    """Folha-baseline por pessoa (salário/encargos + taxas efetivas INSS/FGTS), do snapshot JSON.
    A folha é ~estável mês a mês; a comissão/GBV vêm do vw_comissao por mês."""
    path = os.path.join(os.path.dirname(__file__), "data", "assist_vs_vendedor.json")
    with open(path, encoding="utf-8") as fh:
        snap = json.load(fh)
    base = {}
    DSR = float(snap.get("dsr_fact", 0.20))
    for p in snap.get("people", []):
        com0 = float(p.get("comissao") or 0); dsr0 = com0 * DSR; b0 = com0 + dsr0
        # salary-only items (exclude commission-related: "Comissão", "DSR s/ comissão", "INSS s/ comissão", "FGTS s/ comissão")
        bd_folha = [b for b in p.get("breakdown", []) if "comiss" not in b["label"].lower()]
        base[p["email"].lower()] = {
            "nome": p["nome"], "grupo": p["grupo"], "job": p.get("job", ""), "rel": p.get("rel", ""),
            "sal_cheio": p.get("sal_cheio", 0), "sal_prorata": p.get("sal_prorata", 0),
            "folha_total": float(p.get("folha_total") or 0),
            "inss_rate": (float(p.get("impacto_inss") or 0) / b0) if b0 else 0.0,
            "fgts_rate": (float(p.get("impacto_fgts") or 0) / b0) if b0 else 0.0,
            "breakdown_folha": bd_folha,
        }
    return base, DSR

def _av_month(mes, base, DSR):
    """Computa rentabilidade por pessoa p/ um mês (comissão+GBV do vw_comissao + folha-baseline)."""
    rows = run_query("""
        SELECT LOWER(vendedor) AS email,
               CAST(COALESCE(vlr_final_comissao,0) AS FLOAT64)            AS com,
               CAST(COALESCE(gbv_churn_descontado_transaction,0) AS FLOAT64) AS gbv,
               CAST(COALESCE(atingimento_meta,0) AS FLOAT64)              AS ating,
               CAST(COALESCE(multiplicador,0) AS FLOAT64)                 AS mult,
               MAX(COALESCE(dias_decorridos,0))                           AS dias,
               MAX(CAST(is_projecao AS INT64))                            AS proj
        FROM `fluency-finance.commission.vw_comissao`
        WHERE DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
        GROUP BY 1,2,3,4,5
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=120)
    hist = {r["email"]: r for r in rows}
    people = []
    for email, b in base.items():
        h = hist.get(email)
        com = float(h["com"]) if h else 0.0
        gbv = float(h["gbv"]) if h else 0.0
        proj = bool(h["proj"]) if h else False
        dias = int(h["dias"]) if (h and h["dias"]) else 30
        # projeção (mês parcial): pró-rata a folha pelos dias decorridos p/ ROI comparável
        ff = (dias / 30.0) if proj else 1.0
        folha = round(b["folha_total"] * ff, 2)
        dsr = round(com * DSR, 2)
        inss = round((com + dsr) * b["inss_rate"], 2)
        fgts = round((com + dsr) * b["fgts_rate"], 2)
        imp = round(dsr + inss + fgts, 2)
        custo = round(folha + com + imp, 2)
        # rebuild breakdown: pro-rate salary items if partial month, then append dynamic commission items
        bd_folha = b.get("breakdown_folha", [])
        if proj and ff != 1.0:
            bd_folha = [{"label": x["label"], "val": round(x["val"] * ff)} for x in bd_folha]
        inss_pct = round(b["inss_rate"] * 100, 1)
        fgts_pct = round(b["fgts_rate"] * 100, 1)
        breakdown = bd_folha + [
            {"label": "Comissão", "val": round(com)},
            {"label": f"DSR s/ comissão ({round(DSR*100)}%)", "val": round(dsr)},
            {"label": f"INSS s/ comissão ({inss_pct}%)", "val": round(inss)},
            {"label": f"FGTS s/ comissão ({fgts_pct}%)", "val": round(fgts)},
        ]
        people.append({
            "nome": b["nome"], "email": email, "grupo": b["grupo"], "job": b["job"], "rel": b["rel"],
            "gbv": round(gbv, 2), "ating": (float(h["ating"]) if h else None), "mult": (float(h["mult"]) if h else None),
            "sal_cheio": b["sal_cheio"], "sal_prorata": (round(b["sal_prorata"]*ff,2) if proj else b["sal_prorata"]),
            "folha_total": folha, "comissao": com, "impacto_dsr": dsr, "impacto_inss": inss,
            "impacto_fgts": fgts, "impactos": imp, "custo_total": custo,
            "roi": round(gbv / custo, 1) if custo else 0.0,
            "roi_com": round(gbv / com, 1) if com else 0.0,
            "com_gbv": round(com / gbv, 4) if gbv else None,
            "is_projecao": proj, "dias_decorridos": dias, "breakdown": breakdown,
        })
    return people

def _av_grupos(people):
    grp = {}
    for r in people:
        k = r["grupo"].lower()
        g = grp.setdefault(k, {"grupo": r["grupo"], "n": 0, "n_com_venda": 0, "gbv": 0.0, "folha": 0.0,
                               "comissao": 0.0, "custo": 0.0, "_ga": 0.0, "_ca": 0.0, "custo_ocioso": 0.0})
        g["n"] += 1; g["gbv"] += r["gbv"]; g["folha"] += r["folha_total"]; g["comissao"] += r["comissao"]; g["custo"] += r["custo_total"]
        if r["gbv"] > 0:
            g["n_com_venda"] += 1; g["_ga"] += r["gbv"]; g["_ca"] += r["comissao"]
        else:
            g["custo_ocioso"] += r["custo_total"]
    for g in grp.values():
        g["ativacao"] = round(g["n_com_venda"]/g["n"], 4) if g["n"] else 0
        g["roi_com"] = round(g["gbv"]/g["comissao"], 1) if g["comissao"] else 0.0
        g["roi_com_ativos"] = round(g["_ga"]/g["_ca"], 1) if g["_ca"] else 0.0
        g["com_gbv"] = round(g["comissao"]/g["gbv"], 4) if g["gbv"] else None
        for k in ["gbv", "folha", "comissao", "custo", "custo_ocioso"]:
            g[k] = round(g[k], 2)
        del g["_ga"]; del g["_ca"]
    return {"vendedor": grp.get("vendedor"), "assistente": grp.get("assistente")}

@app.route("/api/assist-vs-vendedor")
@login_required
def api_assist_vs_vendedor():
    if _get_role_data()["role"] != "master":
        return jsonify({"error": "forbidden"}), 403
    mes = resolve_month()
    try:
        base, DSR = _av_baseline()
    except FileNotFoundError:
        return jsonify({"error": "baseline não gerado"}), 404
    people = _av_month(mes, base, DSR)
    grupos = _av_grupos(people)
    # histórico p/ os gráficos: cada mês com apuração (vendedor+assistente) → grupos + top5/bottom5 por ROI
    from datetime import date as _date
    md = _date.fromisoformat(mes)
    hist_meses = []
    for off in range(5, -1, -1):     # últimos ~6 meses até o selecionado
        yy = md.year; mm = md.month - off
        while mm <= 0: mm += 12; yy -= 1
        hist_meses.append(f"{yy:04d}-{mm:02d}-01")
    historico = []
    agg = {}   # email -> {nome, grupo, roi_sum, n} p/ ranking histórico médio
    for hm in hist_meses:
        pe = _av_month(hm, base, DSR)
        com_total = sum(p["comissao"] for p in pe)
        if com_total <= 0:           # mês sem apuração (ex.: jan-mar zerados) → ignora
            continue
        gp = _av_grupos(pe)
        ativos = [p for p in pe if p["gbv"] > 0 and p["custo_total"] > 0]
        ativos.sort(key=lambda x: x["roi"], reverse=True)
        slim = lambda p: {"nome": p["nome"], "grupo": p["grupo"], "roi": p["roi"], "gbv": p["gbv"], "comissao": p["comissao"]}
        historico.append({
            "mes": hm[:7], "mes_label": _MESES_PT[int(hm[5:7]) - 1][:3] + "/" + hm[2:4],
            "grupos": gp, "top5": [slim(p) for p in ativos[:5]], "bottom5": [slim(p) for p in ativos[-5:][::-1]],
        })
        for p in ativos:
            a = agg.setdefault(p["email"], {"nome": p["nome"], "grupo": p["grupo"], "roi_sum": 0.0, "n": 0})
            a["roi_sum"] += p["roi"]; a["n"] += 1
    ranking_hist = sorted(
        [{"nome": a["nome"], "grupo": a["grupo"], "roi": round(a["roi_sum"]/a["n"], 1), "meses": a["n"]}
         for a in agg.values() if a["n"] > 0],
        key=lambda x: x["roi"], reverse=True)
    return jsonify({
        "mes": mes[:7], "mes_label": _MESES_PT[md.month - 1] + " " + str(md.year),
        "dsr_fact": DSR, "people": people, "grupos": grupos, "historico": historico, "ranking_hist": ranking_hist,
        "fonte": "Comissão+GBV = vw_comissao por mês (Jun = projeção; folha pró-rata nos dias). Folha = baseline Supabase. Impacto = DSR 20% + INSS/FGTS (taxa efetiva por pessoa).",
    })


# ── API: B2B · PARCERIA (F.Corporate) — por transação, isolado da hierarquia B2C
# Regra (Raphael, 2026-07-06): comissão = SUM(purchase_amount) × 3%, atribuída à
# vendedora "dona" da empresa (Carteira, aba AUX da planilha Comissionamento B2C),
# via join tracking_source (tag "[F]_Empresa") x commission.b2b_carteira_map.
# ⚠️ "Parceria" é UM dos modelos de comissionamento B2B (Raphael, 2026-07-06) —
# haverá modelos separados para "Convênio" e "Subsídio Parcial", cada um com sua
# própria regra/rota/aba. NÃO misturar. NÃO usar tracking_source_sck pra atribuir
# vendedora — é o e-mail de quem registrou a venda, não necessariamente a dona da
# conta. NÃO é GBV, NÃO tem vendedor/TL/coordenador B2C associado. Ver skill
# comissao-b2b pro racional completo e pipeline/load_b2b_carteira_map.py (refresh
# do de-para).
# Retorna o grão detalhado (mês × empresa × vendedora, 2026-07-06) — os filtros de
# mês/vendedora/empresa do dashboard e as agregações (por vendedora, por empresa)
# são todos client-side em cima dessa lista, o volume é pequeno (~2.5k transações).
@app.route("/api/b2b-parceria")
@login_required
def api_b2b_parceria():
    if _get_role_data()["role"] not in ("master", "gestor"):
        return jsonify({"error": "forbidden"}), 403
    rows = run_query("""
        WITH src AS (
          SELECT
            FORMAT_DATE('%Y-%m', DATE(approved_date_brt)) AS mes,
            LOWER(TRIM(tracking_source)) AS tag_norm,
            purchase_amount
          FROM `fluency-silver.hotmart.transactions`
          WHERE LOWER(product_name) LIKE '%f.corporate%'
            AND purchase_status IN ('COMPLETE', 'APPROVED')
        )
        SELECT
          src.mes                    AS mes,
          COALESCE(m.tag_empresa, 'Não mapeado') AS empresa,
          CASE
            WHEN m.carteira IS NULL THEN 'Não mapeado'
            WHEN m.carteira IN ('Avulso', '-') THEN 'Sem carteira'
            ELSE m.carteira
          END                        AS vendedora,
          COUNT(*)                   AS qtd_transacoes,
          SUM(src.purchase_amount)   AS total_purchase_amount
        FROM src
        LEFT JOIN `fluency-finance.commission.b2b_carteira_map` m
          ON m.tag_empresa_norm = src.tag_norm
        GROUP BY 1, 2, 3
        ORDER BY 1 DESC, 5 DESC
    """, [], cache_ttl=90)
    linhas = []
    for r in rows:
        amt = float(r["total_purchase_amount"] or 0)
        qtd = int(r["qtd_transacoes"] or 0)
        linhas.append({
            "mes": r["mes"],
            "empresa": r["empresa"],
            "vendedora": r["vendedora"],
            "qtd_transacoes": qtd,
            "total_purchase_amount": round(amt, 2),
            "comissao": round(amt * 0.03, 2),
        })
    return jsonify({
        "modelo": "Parceria",
        "linhas": linhas,
        "fonte": "Modelo B2B · Parceria (b2b2c) — fluency-silver.hotmart.transactions (product_name LIKE '%f.corporate%', status COMPLETE/APPROVED) × commission.b2b_carteira_map (de-para AUX) por tracking_source. Comissão = purchase_amount × 3%. \"Sem carteira\" = tag existe na AUX mas sem dono definido (Avulso/-). \"Não mapeado\" = tag sem correspondência na AUX (ou transação sem tracking_source). Grão: mês × empresa (tag_empresa) × vendedora (carteira). Outros modelos B2B (Convênio, Subsídio Parcial) têm regra e endpoint próprios.",
    })


# ── API: B2B (Entradas, AC="B2B") — Convênio + Portal de Benefícios + Fluency
# Pass + Subsídio Parcial, isolado do modelo b2b2c/Parceria ────────────────────
# Regra (Raphael, 2026-07-06): comissão = SUM(coluna L "Valor de compra com
# impostos") × 6%, por vendedora (coluna AI "Carteira"), quebrado por Contrato
# (coluna AD). Corte de mês = coluna F ("Confirmação do pagamento"). Fonte: aba
# Entradas da planilha Comissionamento B2C — snapshot em
# commission.b2b_entradas_snapshot (pipeline/load_b2b_entradas.py).
# ⚠️ Essa planilha é editada ao vivo pelo time — o snapshot reflete o momento em
# que o loader rodou por último, não é live. ⚠️ "Aguardando validação" = linhas
# com #REF! na coluna AI (fórmula quebrada na sheet) — EXCLUÍDO dos totais
# oficiais até o Raphael corrigir e o snapshot ser refeito (decisão 2026-07-06).
# "Sem carteira" = coluna AI vazia ou "-" — ENTRA no total (decisão 2026-07-06).
@app.route("/api/b2b")
@login_required
def api_b2b():
    if _get_role_data()["role"] not in ("master", "gestor"):
        return jsonify({"error": "forbidden"}), 403
    rows = run_query("""
        SELECT mes, empresa, contrato, vendedora, COUNT(*) AS qtd, SUM(valor) AS total
        FROM `fluency-finance.commission.b2b_entradas_snapshot`
        GROUP BY 1, 2, 3, 4
        ORDER BY 1 DESC, 6 DESC
    """, [], cache_ttl=90)
    # Exceção Raphael 2026-07-07: contas "Unico Skill" (qualquer grafia, ex.
    # "Unico Skill (SKILLHUB)"/"Unico Skill (Skillhub)") comissionam a 3%, não
    # os 6% padrão do modelo B2B/Convênio.
    linhas = []
    for r in rows:
        empresa = r["empresa"] or ""
        taxa = 0.03 if empresa.strip().lower().startswith("unico skill") else 0.06
        linhas.append({
            "mes": r["mes"],
            "empresa": r["empresa"],
            "contrato": r["contrato"],
            "vendedora": r["vendedora"],
            "qtd_transacoes": int(r["qtd"] or 0),
            "total_valor": round(float(r["total"] or 0), 2),
            "comissao": round(float(r["total"] or 0) * taxa, 2),
        })
    return jsonify({
        "modelo": "B2B",
        "linhas": linhas,
        "fonte": "Modelo B2B — aba Entradas (planilha Comissionamento B2C), AC=\"B2B\" (Convênio + Portal de Benefícios + Fluency Pass + Subsídio Parcial). Comissão = coluna L (Valor de compra com impostos) × 6%, por vendedora (Carteira) — EXCEÇÃO: contas \"Unico Skill\" comissionam a 3%. Snapshot estático (pipeline/load_b2b_entradas.py) — a sheet é editada ao vivo. \"Sem carteira\" entra no total; \"Aguardando validação\" (#REF! na sheet, vendedora='Aguardando validação') fica FORA do total oficial até correção. Grão: mês × empresa (Comprador) × contrato × vendedora.",
    })


# ── API: transactions ─────────────────────────────────────────────────────────

# Taxas por modelo = MESMAS do pipeline calc_comissao.py (NÃO a regra_formas_pagamento,
# que está com a taxa da Tacyana bugada 30%/10%). OTE → sem comissão por transação.
_RATE_ANALISTA   = {"a vista": 0.10, "parcelado": 0.04,  "inteligente": 0.013}
_RATE_DELUCHI    = {"a vista": 0.08, "parcelado": 0.04,  "inteligente": 0.0225}  # Recuperação
_RATE_TL_VANESSA = {"a vista": 0.04, "parcelado": 0.015, "inteligente": 0.005}
_RATE_TL_TACYANA = {"a vista": 0.03, "parcelado": 0.01,  "inteligente": 0.005}

def _rates_for_vendedor(email: str, cargo: str):
    """Taxas por forma de pagamento do modelo (comissão por transação, PRÉ-multiplicador).
    OTE (assistente, TL Novo I = Ana/Matheus, coordenador) → None ('—')."""
    e = (email or "").lower()
    if e == "vanessa.lopes@fluencyacademy.io":
        return _RATE_TL_VANESSA
    if e == "tacyana.bueno@fluencyacademy.io":
        return _RATE_TL_TACYANA
    if e == "ana.deluchi@fluencyacademy.io":
        return _RATE_DELUCHI
    if (cargo or "").lower() == "vendedor":   # Analista (% por modalidade)
        return _RATE_ANALISTA
    return None

@app.route("/api/transactions")
@login_required
def api_transactions():
    return jsonify(_compute_transactions(resolve_target(effective_email()), resolve_month()))

def _compute_transactions(target, mes):
    """Lista de transações processadas do colaborador no mês (mesma fonte da aba + PDF de fechamento)."""
    # Detalhe transação a transação direto da base nova (obt + dm_orders) — a antiga
    # commission.vendedores parou de ser carregada em 04/2026. is_churn_tx e recebível
    # seguem a mesma regra do CTE da view_comissao_vendedores.
    sql = """
        WITH vendas AS (
          SELECT contract_id, transaction_id, vendedor, gbv, modality_payment, tracking_source_sck,
                 ANY_VALUE(contract_email) AS contract_email
          FROM `fluency-gold.conversion.obt_conversions`
          WHERE transaction_id IS NOT NULL AND contract_id IS NOT NULL
          GROUP BY contract_id, transaction_id, vendedor, gbv, modality_payment, tracking_source_sck
        )
        SELECT
          t.transaction_id,
          t.transaction_status,
          DATE(t.contract_created_at_brt_timestamp)               AS data_contrato,
          v.modality_payment                                      AS modality_payment,
          v.contract_email                                        AS cliente_email,
          CAST(ROUND(v.gbv, 2)          AS NUMERIC)               AS gbv,
          CAST(t.transaction_amount     AS NUMERIC)               AS parcela,
          IF(t.transaction_status IS DISTINCT FROM 'paid'
             AND DATE_DIFF(DATE(t.contract_created_at_brt_timestamp),
                           DATE(t.transaction_confirmation_purchase_at_brt_timestamp), DAY) < 8,
             TRUE, FALSE)                                          AS is_churn_tx,
          CASE WHEN t.transaction_status = 'paid'                          THEN 'recebido'
               WHEN t.transaction_status IN ('refunded','chargeback','dispute') THEN 'cancelado'
               ELSE 'a_receber' END                               AS recebivel
        FROM vendas v
        JOIN `fluency-gold.sales.dm_contracts_orders_transaction` t
          ON v.transaction_id = t.transaction_id
        -- double-check: atribui por `vendedor` OU pelo e-mail no `tracking_source_sck`
        -- (vendas de esteira vêm com vendedor NULL e o e-mail no tracking)
        WHERE (LOWER(v.vendedor) = LOWER(@email)
               OR LOWER(v.tracking_source_sck) LIKE CONCAT('%', LOWER(@email), '%'))
          AND DATE_TRUNC(DATE(t.contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
        ORDER BY t.contract_created_at_brt_timestamp DESC
    """
    rows = run_query(sql, [
        bigquery.ScalarQueryParameter("email", "STRING", target),
        bigquery.ScalarQueryParameter("mes",   "DATE",   mes),
    ])
    # Comissão por transação:
    #   • % models (Analista/TL%): GBV líq × taxa da forma × multiplicador (reflete à vista/parcelado/inteligente)
    #   • OTE (assistente/TL-novo/coord): não há taxa transacional → aloca o vlr_final proporcional ao GBV líq
    # Em ambos os casos a coluna SOMA o vlr_final do mês (snapshot comissao_historica).
    cargo_t = _load_vmap().get(target.lower(), {}).get("cargo", "")
    rates   = _rates_for_vendedor(target, cargo_t)
    snap = run_query("""
        SELECT CAST(COALESCE(multiplicador,1)      AS FLOAT64) AS mult,
               CAST(COALESCE(atingimento_meta,0)    AS FLOAT64) AS ating,
               CAST(COALESCE(vlr_final_comissao,0)  AS FLOAT64) AS vlr_final
        FROM `fluency-finance.commission.vw_comissao`
        WHERE LOWER(vendedor)=LOWER(@email)
          AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH)=DATE(@mes) LIMIT 1
    """, [bigquery.ScalarQueryParameter("email", "STRING", target),
          bigquery.ScalarQueryParameter("mes",   "DATE",   mes)], cache_ttl=120)
    mult      = float(snap[0]["mult"])      if snap else 1.0
    _snap_ating = float(snap[0].get("ating", 0) or 0) if snap else 0.0
    vlr_final = _corrige_vlr_ote(
        float(snap[0]["vlr_final"]) if snap else 0.0,
        _modelo_comissao(target, cargo_t), _snap_ating, mult)
    # base da alocação OTE = Σ GBV líquido (pós-churn) das transações de sistema
    total_liq = sum((0.0 if bool(r.get("is_churn_tx"))
                     else (float(r["gbv"]) if r.get("gbv") is not None else 0.0)) for r in rows)
    ote_rate  = (vlr_final / total_liq) if (not rates and total_liq > 0) else None
    for r in rows:
        churn = bool(r.get("is_churn_tx"))
        gbv   = float(r["gbv"]) if r.get("gbv") is not None else 0.0
        liq   = 0.0 if churn else gbv
        modality = (r.get("modality_payment") or "").strip().lower()
        r["data_contrato"] = str(r["data_contrato"]) if r["data_contrato"] else None
        r["gbv"]           = gbv
        r["parcela"]       = float(r["parcela"]) if r.get("parcela") is not None else 0.0
        r["gbv_liquido"]   = liq
        r["is_churn"]      = 1 if churn else 0
        if rates and modality in rates:
            r["comissao"] = round(liq * rates[modality] * mult, 2)   # % model: taxa da forma × mult
        elif ote_rate is not None:
            r["comissao"] = round(liq * ote_rate, 2)                 # OTE: proporcional ao GBV
        else:
            r["comissao"] = None
        r["origem"]        = "sistema"
        r["aprovacao_hp"]  = None   # transação de sistema não passa por aprovação de HP
        r["forma_pagamento"] = (r.get("modality_payment") or "").strip().lower() or None
        r.pop("is_churn_tx", None)
        r.pop("modality_payment", None)
    # Extras (HP) entram como linhas no dash (não na fonte). Mostra TODOS (pendente/aprovado/
    # rejeitado) com a dupla aprovação TL+coord — só os EFETIVOS (ambos aprovados) somam no GBV
    # (isso é feito em /api/summary via _approved_extras).
    for e in _extras_for_display(target, mes):
        gbv = float(e["gbv"]) if e.get("gbv") is not None else 0.0
        is_churn = int(e["is_churn"]) if e.get("is_churn") is not None else 0
        st_tl, st_coord = e.get("status_tl") or "pendente", e.get("status_coord") or "pendente"
        efetivo = (st_tl == "aprovado" and st_coord == "aprovado")
        rows.append({
            "transaction_id":     e.get("transaction_id"),
            "transaction_status": "aprovado" if efetivo else "pendente",
            "data_contrato":      str(e["created_at"])[:10] if e.get("created_at") else None,
            "gbv":                gbv,
            "parcela":            gbv,
            "gbv_liquido":        0.0 if is_churn else gbv,
            "is_churn":           is_churn,
            "comissao":           None,
            "recebivel":          "a_receber",
            "origem":             "extra",
            "forma_pagamento":    (e.get("modality_payment") or "").strip().lower() or None,
            "cliente_email":      e.get("cliente_email"),
            "aprovacao_hp":       {"id": e.get("id"), "status_tl": st_tl, "status_coord": st_coord},
        })
    return rows

# ── API: approvals (read) ─────────────────────────────────────────────────────

@app.route("/api/approvals")
@login_required
def api_approvals():
    target = resolve_target(effective_email())
    mes    = resolve_month()
    sql = """
        SELECT transaction_id, status, comment, approved_at
        FROM `fluency-finance.commission.approval_vendedores`
        WHERE LOWER(vendedor) = LOWER(@email)
          AND competencia = DATE(@mes)
    """
    rows = run_query(sql, [
        bigquery.ScalarQueryParameter("email", "STRING", target),
        bigquery.ScalarQueryParameter("mes",   "DATE",   mes),
    ])
    result = {}
    for r in rows:
        result[r["transaction_id"]] = {
            "status":  r["status"],
            "comment": r["comment"] or "",
            "at":      str(r["approved_at"]),
        }
    return jsonify(result)

# ── API: save approval ────────────────────────────────────────────────────────

@app.route("/api/approvals", methods=["POST"])
@login_required
def api_save_approval():
    # Só o PRÓPRIO vendedor/assistente (role efetivo 'vendedor') aprova as próprias transações.
    # Bloqueia gestor/tl/coord/master/people_ops e qualquer simulação ("ver como").
    if _get_role_data()["role"] != "vendedor" or _view_as():
        return jsonify({"error": "forbidden"}), 403
    email = session["email"]
    body  = request.get_json()
    tx_id   = body.get("transaction_id")
    status  = body.get("status")
    comment = body.get("comment", "")
    mes     = body.get("mes", current_month_brt())

    if not tx_id or not status:
        return jsonify({"error": "transaction_id e status são obrigatórios"}), 400

    sql = """
        MERGE `fluency-finance.commission.approval_vendedores` T
        USING (
          SELECT
            @tx_id   AS transaction_id,
            @email   AS vendedor,
            @status  AS status,
            @comment AS comment,
            DATE(@mes) AS competencia,
            CURRENT_TIMESTAMP() AS approved_at,
            @email   AS approved_by
        ) S
        ON T.transaction_id = S.transaction_id
           AND LOWER(T.vendedor) = LOWER(S.vendedor)
        WHEN MATCHED THEN
          UPDATE SET status=S.status, comment=S.comment, approved_at=S.approved_at
        WHEN NOT MATCHED THEN
          INSERT (transaction_id, vendedor, status, comment, competencia, approved_at, approved_by)
          VALUES (S.transaction_id, S.vendedor, S.status, S.comment, S.competencia, S.approved_at, S.approved_by)
    """
    bq.query(sql, job_config=bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("tx_id",   "STRING", tx_id),
        bigquery.ScalarQueryParameter("email",   "STRING", email),
        bigquery.ScalarQueryParameter("status",  "STRING", status),
        bigquery.ScalarQueryParameter("comment", "STRING", comment),
        bigquery.ScalarQueryParameter("mes",     "DATE",   mes),
    ])).result()
    return jsonify({"ok": True})

# ── API: bulk approve ─────────────────────────────────────────────────────────

@app.route("/api/approvals/bulk", methods=["POST"])
@login_required
def api_bulk_approve():
    if _get_role_data()["role"] in READONLY_ROLES or _view_as():
        return jsonify({"error": "forbidden"}), 403
    email = session["email"]
    body  = request.get_json()
    ids   = body.get("transaction_ids", [])
    mes   = body.get("mes", current_month_brt())
    if not ids:
        return jsonify({"ok": True, "count": 0})

    for tx_id in ids:
        sql = """
            MERGE `fluency-finance.commission.approval_vendedores` T
            USING (
              SELECT
                @tx_id     AS transaction_id,
                @email     AS vendedor,
                'aprovado' AS status,
                ''         AS comment,
                DATE(@mes) AS competencia,
                CURRENT_TIMESTAMP() AS approved_at,
                @email     AS approved_by
            ) S
            ON T.transaction_id = S.transaction_id
               AND LOWER(T.vendedor) = LOWER(S.vendedor)
            WHEN MATCHED THEN
              UPDATE SET status='aprovado', approved_at=S.approved_at
            WHEN NOT MATCHED THEN
              INSERT (transaction_id, vendedor, status, comment, competencia, approved_at, approved_by)
              VALUES (S.transaction_id, S.vendedor, S.status, S.comment, S.competencia, S.approved_at, S.approved_by)
        """
        bq.query(sql, job_config=bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("tx_id", "STRING", tx_id),
            bigquery.ScalarQueryParameter("email", "STRING", email),
            bigquery.ScalarQueryParameter("mes",   "DATE",   mes),
        ])).result()

    return jsonify({"ok": True, "count": len(ids)})

# ── API: sign-off do vendedor ("Estou de acordo") ─────────────────────────────

def _mes_label(mes: str) -> str:
    d = date.fromisoformat(mes)
    return f"{_MONTH_ABBR[d.month-1]}/{d.year}"

def _mes_extenso(mes: str) -> str:
    d = date.fromisoformat(mes)
    return f"{_MESES_PT[d.month-1]}/{d.year}"

# ── Aviso People Ops: dispara quando 100% do roster elegível aceitou ──────────
def _eligible_signoff_roster(mes: str) -> set:
    rows = run_query("""
        SELECT LOWER(email_vendedor) AS email
        FROM `fluency-finance.commission.hierarquia_comercial`
        WHERE mes_venda = DATE(@mes) AND cargo IN ('Vendedor','Assistente')
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)])
    inactive = _inactive_emails_for_month(mes)
    return {r["email"] for r in rows if r["email"] and (not inactive or r["email"] not in inactive)}

def _notify_people_ops_if_complete(mes: str) -> None:
    if _view_as():
        return
    # idempotência: verifica se já enviamos o aviso neste mês
    already = run_query("""
        SELECT 1 FROM `fluency-finance.commission.signoff_vendedores`
        WHERE competencia = DATE(@mes) AND aviso_people_ops_at IS NOT NULL LIMIT 1
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)], cache_ttl=0)
    if already:
        return
    # Gatilho: aprovador final (Yani) assinou
    aprov_signed = run_query("""
        SELECT signed_at FROM `fluency-finance.commission.signoff_vendedores`
        WHERE LOWER(vendedor) = @aprov AND competencia = DATE(@mes) LIMIT 1
    """, [bigquery.ScalarQueryParameter("aprov", "STRING", APROVADOR_FINAL_EMAIL.lower()),
          bigquery.ScalarQueryParameter("mes",   "DATE",   mes)], cache_ttl=0)
    if not aprov_signed:
        return
    mes_ext = _mes_extenso(mes)
    aprov_nome = _name_from_email(APROVADOR_FINAL_EMAIL)
    now_brt = (datetime.now(timezone.utc) - timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")
    html = (
        f"<p>Olá,</p>"
        f"<p><strong>{aprov_nome}</strong> aprovou o fechamento de comissões de "
        f"<strong>{mes_ext}</strong> em {now_brt}.</p>"
        f"<p>O fechamento está consolidado e <strong>liberado para processamento na folha</strong>.</p>"
        f"<p><a href='https://commission-dashboard-m7r6fovrsa-uc.a.run.app/dashboard'>"
        f"Acessar dashboard de comissões</a></p>"
        f"<p>Atenciosamente,<br><strong>Finance — Fluency Academy</strong></p>"
    )
    send_mail(PEOPLE_OPS_NOTIFY_EMAILS,
              f"Comissões de {mes_ext} aprovadas — pode processar",
              html)
    # marca idempotência na linha do aprovador
    run_query("""
        UPDATE `fluency-finance.commission.signoff_vendedores`
        SET aviso_people_ops_at = CURRENT_TIMESTAMP()
        WHERE LOWER(vendedor) = @aprov AND competencia = DATE(@mes)
    """, [bigquery.ScalarQueryParameter("aprov", "STRING", APROVADOR_FINAL_EMAIL.lower()),
          bigquery.ScalarQueryParameter("mes",   "DATE",   mes)])

@app.route("/api/signoff", methods=["GET"])
@login_required
def api_signoff_get():
    role = _get_role_data()["role"]
    # Papéis privilegiados podem consultar o aceite de qualquer vendedor via ?vendedor=
    if role in ("tl", "gestor", "master", "people_ops"):
        v = request.args.get("vendedor", "").strip().lower()
        target = v if v and v != "todos" else effective_email()
    else:
        target = effective_email()
    mes = resolve_month()
    rows = run_query("""
        SELECT signed_at FROM `fluency-finance.commission.signoff_vendedores`
        WHERE LOWER(vendedor) = LOWER(@v) AND competencia = DATE(@mes) LIMIT 1
    """, [
        bigquery.ScalarQueryParameter("v",   "STRING", target),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])
    if _is_newcomer_assist(target, mes):
        return jsonify({"signed": False, "at": None})
    at = str(rows[0]["signed_at"]) if rows and rows[0].get("signed_at") else None
    return jsonify({"signed": bool(rows), "at": at})

def _build_team_pdf(tl_nome: str, coord_nome: str | None, label: str, members: list) -> bytes:
    """PDF resumo do fechamento completo do time: vendedor | TL | comissão."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=14*mm, bottomMargin=14*mm,
                            title=f"Fechamento time {tl_nome} {label}")
    ss = getSampleStyleSheet()
    PURPLE = colors.HexColor("#7B5CF6")
    h  = ParagraphStyle("h",  parent=ss["Title"],  fontSize=15, spaceAfter=2, textColor=PURPLE)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=9,  textColor=colors.HexColor("#64748b"))
    el = []
    el.append(Paragraph("Fluency · Fechamento Completo de Comissões", h))
    coord_line = f" · Coordenador: {coord_nome}" if coord_nome else ""
    el.append(Paragraph(f"Time: <b>{tl_nome}</b>{coord_line} · Competência: <b>{label}</b>", sub))
    el.append(Spacer(1, 10))
    # Tabela de membros
    head = ["Vendedor", "TL", "Comissão Final"]
    data = [head] + [
        [m["nome"], tl_nome, _brl(m["comissao"])]
        for m in sorted(members, key=lambda x: x["nome"])
    ]
    total_com = sum(m["comissao"] for m in members)
    data.append(["TOTAL", "", _brl(total_com)])
    col_w = [90*mm, 60*mm, 36*mm]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  PURPLE),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("ALIGN",         (2, 1), (2, -1),  "RIGHT"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -2), [colors.white, colors.HexColor("#f7f8fb")]),
        ("FONTNAME",      (0, -1),(-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",     (2, -1),(2, -1),  PURPLE),
        ("LINEABOVE",     (0, -1),(-1, -1), 0.8, PURPLE),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    el.append(tbl)
    el.append(Spacer(1, 8))
    el.append(Paragraph(
        f"<i>Todos os {len(members)} colaboradores confirmaram o aceite de comissões de {label}.</i>",
        ParagraphStyle("note", parent=ss["Normal"], fontSize=8, textColor=colors.HexColor("#94a3b8"))
    ))
    doc.build(el)
    return buf.getvalue()


def _notify_if_team_complete(tl: str, mes: str, label: str, chain: dict):
    """Envia Template C (fechamento completo) se todos os vendedores do time do TL assinaram."""
    if not tl:
        return
    rows = run_query("""
        SELECT LOWER(hc.email_vendedor) AS email,
               sv.signed_at IS NOT NULL  AS assinou
        FROM `fluency-finance.commission.hierarquia_comercial` hc
        LEFT JOIN `fluency-finance.commission.signoff_vendedores` sv
          ON LOWER(hc.email_vendedor) = LOWER(sv.vendedor)
         AND sv.competencia = DATE(@mes)
        WHERE LOWER(hc.gestor) = LOWER(@tl) AND hc.mes_venda = DATE(@mes)
    """, [bigquery.ScalarQueryParameter("tl", "STRING", tl),
          bigquery.ScalarQueryParameter("mes", "DATE", mes)])
    if not rows:
        return
    total = len(rows)
    assinados = sum(1 for r in rows if r["assinou"])
    if total == 0 or assinados < total:
        return

    # Busca comissão de cada membro no mês
    emails_time = [r["email"] for r in rows]
    com_rows = run_query("""
        SELECT LOWER(vendedor) AS email,
               MAX(CAST(COALESCE(vlr_final_comissao,0) AS FLOAT64)) AS comissao
        FROM `fluency-finance.commission.vw_comissao`
        WHERE DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
          AND LOWER(vendedor) IN UNNEST(@emails)
        GROUP BY 1
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes),
          bigquery.ArrayQueryParameter("emails", "STRING", emails_time)])
    com_map = {r["email"]: float(r["comissao"] or 0) for r in com_rows}

    tl_nome   = _name_from_email(tl)
    coord     = chain.get("coord", "")
    # coord_nome só aparece se for diferente do TL
    coord_nome = _name_from_email(coord) if coord and coord.lower() != tl.lower() else None

    members = [{"nome": _name_from_email(r["email"]), "email": r["email"],
                "comissao": com_map.get(r["email"], 0.0)} for r in rows]

    # Gera PDF e faz upload para GCS
    pdf_url = None
    try:
        pdf_bytes = _build_team_pdf(tl_nome, coord_nome, label, members)
        pdf_url = _upload_pdf_gcs(pdf_bytes, f"time-{tl.split('@')[0]}", mes)
    except Exception as _e:
        app.logger.warning("PDF fechamento completo falhou: %s", _e)

    pdf_link_html = (
        f"<p><a href='{pdf_url}' style='display:inline-block;padding:8px 16px;background:#7B5CF6;"
        f"color:white;text-decoration:none;border-radius:6px;font-weight:600;font-size:14px'>"
        f"⬇ Baixar resumo do time (PDF)</a>"
        f"<span style='color:#888;font-size:11px;margin-left:8px'>(link válido por 7 dias)</span></p>"
    ) if pdf_url else ""

    nomes_lista = "".join(
        f"<li>{m['nome']} — {_brl(m['comissao'])}</li>"
        for m in sorted(members, key=lambda x: x["nome"])
    )
    html_completo = (
        f"<p>Olá,</p>"
        f"<p>O time de <b>{tl_nome}</b> concluiu o <b>fechamento de comissões de {label}</b>. "
        f"Todos os {total} colaboradores confirmaram o aceite:</p>"
        f"<ul style='font-size:14px;line-height:1.8'>{nomes_lista}</ul>"
        f"{pdf_link_html}"
        f"<p>Acesse o <a href='https://commission-dashboard-m7r6fovrsa-uc.a.run.app/dashboard'>dashboard</a> "
        f"para ver o status completo do time.</p>"
        f"<p>— Finance · Fluency Academy</p>"
    )
    notif_to = list(dict.fromkeys(e for e in [tl, chain["coord"]] + SIGNOFF_CC_FIXO if e))
    send_mail(notif_to, f"Fechamento completo — Time {tl_nome} · {label}", html_completo)


@app.route("/api/signoff", methods=["POST"])
@login_required
def api_signoff_post():
    role = _get_role_data()["role"]
    if role in READONLY_ROLES or _view_as():
        return jsonify({"error": "forbidden"}), 403
    vend = session["email"].lower()
    mes  = request.get_json(silent=True).get("mes") if request.is_json else None
    mes  = (mes + "-01") if (mes and len(mes) == 7) else current_month_brt()

    # Aprovador final: salva o aceite e dispara direto a notificação People Ops.
    if role == "aprovador":
        bq.query("""
            MERGE `fluency-finance.commission.signoff_vendedores` T
            USING (SELECT @v AS vendedor, DATE(@mes) AS competencia,
                          CURRENT_TIMESTAMP() AS signed_at, @v AS signed_by) S
            ON LOWER(T.vendedor) = LOWER(S.vendedor) AND T.competencia = S.competencia
            WHEN MATCHED THEN UPDATE SET signed_at = S.signed_at, signed_by = S.signed_by
            WHEN NOT MATCHED THEN INSERT (vendedor, competencia, signed_at, signed_by)
              VALUES (S.vendedor, S.competencia, S.signed_at, S.signed_by)
        """, job_config=bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("v",   "STRING", vend),
            bigquery.ScalarQueryParameter("mes", "DATE",   mes),
        ])).result()
        try:
            _notify_people_ops_if_complete(mes)
        except Exception as _e:
            app.logger.warning("aviso people_ops (aprovador) falhou: %s", _e)
        return jsonify({"ok": True})

    # Vendedores/assistentes só podem assinar após todos os HPs estarem resolvidos.
    if role == "vendedor":
        hp_rows = run_query("""
            SELECT COUNT(*) AS n
            FROM `fluency-finance.commission.extras_vendedores`
            WHERE LOWER(vendedor) = LOWER(@v)
              AND competencia = DATE(@mes)
              AND COALESCE(status_tl, '') NOT IN ('rejeitado')
              AND COALESCE(status_coord, '') NOT IN ('aprovado', 'rejeitado')
        """, [bigquery.ScalarQueryParameter("v",   "STRING", vend),
              bigquery.ScalarQueryParameter("mes", "DATE",   mes)])
        pending = int(hp_rows[0]["n"]) if hp_rows else 0
        if pending > 0:
            return jsonify({
                "error": f"Você tem {pending} HP(s) pendentes de aprovação — aguarde o TL antes de assinar.",
                "code":  "hp_pendente",
                "count": pending,
            }), 409

    bq.query("""
        MERGE `fluency-finance.commission.signoff_vendedores` T
        USING (SELECT @v AS vendedor, DATE(@mes) AS competencia,
                      CURRENT_TIMESTAMP() AS signed_at, @v AS signed_by) S
        ON LOWER(T.vendedor) = LOWER(S.vendedor) AND T.competencia = S.competencia
        WHEN MATCHED THEN UPDATE SET signed_at = S.signed_at, signed_by = S.signed_by
        WHEN NOT MATCHED THEN INSERT (vendedor, competencia, signed_at, signed_by)
          VALUES (S.vendedor, S.competencia, S.signed_at, S.signed_by)
    """, job_config=bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("v",   "STRING", vend),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])).result()

    chain = _commission_chain(vend, mes)
    nome  = _name_from_email(vend)
    label = _mes_label(mes)
    cargo = _load_vmap().get(vend, {}).get("cargo", "")
    now_brt = (datetime.now(timezone.utc) - timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")
    # resumo do snapshot (mesma fonte do dashboard)
    sr = run_query("""
        SELECT CAST(COALESCE(gbv,0)                              AS FLOAT64) AS gbv_bruto,
               CAST(COALESCE(gbv_apenas_churn_transaction,0)     AS FLOAT64) AS churn,
               CAST(COALESCE(qtd_is_churn_transaction,0)         AS INT64)   AS churns,
               CAST(COALESCE(gbv_churn_descontado_transaction,0) AS FLOAT64) AS gbv_liq,
               CAST(COALESCE(atingimento_meta,0)                 AS FLOAT64) AS ating,
               CAST(COALESCE(multiplicador,0)                    AS FLOAT64) AS mult,
               CAST(COALESCE(total_comissao,0)                   AS FLOAT64) AS total_comissao,
               CAST(COALESCE(vlr_final_comissao,0)               AS FLOAT64) AS vlr_final
        FROM `fluency-finance.commission.vw_comissao`
        WHERE LOWER(vendedor)=LOWER(@v)
          AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH)=DATE(@mes) LIMIT 1
    """, [bigquery.ScalarQueryParameter("v",   "STRING", vend),
          bigquery.ScalarQueryParameter("mes", "DATE",   mes)])
    summary = dict(sr[0]) if sr else {
        "gbv_bruto":0,"churn":0,"churns":0,"gbv_liq":0,
        "ating":0,"mult":0,"total_comissao":0,"vlr_final":0,
    }
    # Corrige vlr_final para modelos OTE (pipeline gravou OTE×mult sem atingimento).
    _cargo_sf = _load_vmap().get(vend.lower(), {}).get("cargo", "")
    # Meta e ote_fator do roster (hierarquia_comercial) — igual ao api_summary
    mr = run_query("""
        SELECT CAST(COALESCE(valor_meta,0) AS FLOAT64) AS vm,
               CAST(COALESCE(ote_fator, 1.0) AS FLOAT64) AS ote_fator
        FROM `fluency-finance.commission.hierarquia_comercial`
        WHERE LOWER(email_vendedor)=LOWER(@v) AND mes_venda=DATE(@mes) LIMIT 1
    """, [bigquery.ScalarQueryParameter("v",   "STRING", vend),
          bigquery.ScalarQueryParameter("mes", "DATE",   mes)])
    summary["valor_meta"] = float(mr[0]["vm"]) if mr and mr[0]["vm"] else None
    _ote_fator_sf = float(mr[0]["ote_fator"]) if mr else 1.0
    summary["vlr_final"] = _corrige_vlr_ote(
        float(summary.get("vlr_final") or 0), _modelo_comissao(vend, _cargo_sf),
        float(summary.get("ating") or 0), float(summary.get("mult") or 0), _ote_fator_sf)
    txs = _compute_transactions(vend, mes)
    summary["n"] = len(txs)
    summary["signed_at"] = now_brt
    try:
        pdf = _build_signoff_pdf(nome, vend, label, cargo, chain["tl"], summary, txs)
    except Exception as e:
        app.logger.warning("Falha ao gerar PDF de fechamento: %s", e); pdf = None
    pct  = f"{summary['ating']*100:.2f}".replace(".", ",")
    mult = f"{summary['mult']:.1f}".replace(".", ",")
    resumo_html = (
        f"<table style='border-collapse:collapse;font-size:14px;margin:10px 0'>"
        f"<tr><td style='padding:4px 16px 4px 0;color:#555'>Transações</td><td><b>{summary['n']}</b></td></tr>"
        f"<tr><td style='padding:4px 16px 4px 0;color:#555'>GBV bruto</td><td><b>{_brl(summary['gbv_bruto'])}</b></td></tr>"
        f"<tr><td style='padding:4px 16px 4px 0;color:#555'>Churn</td><td><b>{_brl(summary['churn'])}</b></td></tr>"
        f"<tr><td style='padding:4px 16px 4px 0;color:#555'>GBV líquido</td><td><b>{_brl(summary['gbv_liq'])}</b></td></tr>"
        f"<tr><td style='padding:4px 16px 4px 0;color:#555'>Atingimento</td><td><b>{pct}%</b></td></tr>"
        f"<tr><td style='padding:4px 16px 4px 0;color:#555'>Multiplicador</td><td><b>{mult}×</b></td></tr>"
        f"<tr><td style='padding:4px 16px 4px 0;color:#555;font-weight:700'>Comissão final</td>"
        f"<td><b style='font-size:16px;color:#7B5CF6'>{_brl(summary['vlr_final'])}</b></td></tr>"
        f"</table>"
    )

    # Upload PDF para GCS e gera link de download (7 dias)
    pdf_url = _upload_pdf_gcs(pdf, vend, mes) if pdf else None
    pdf_link_html = (
        f"<p><a href='{pdf_url}' style='display:inline-block;padding:8px 16px;background:#7B5CF6;"
        f"color:white;text-decoration:none;border-radius:6px;font-weight:600;font-size:14px'>"
        f"⬇ Baixar demonstrativo PDF</a>"
        f"<span style='color:#888;font-size:11px;margin-left:8px'>(link válido por 7 dias)</span></p>"
    ) if pdf_url else ""

    # ── Template A: aceite individual — todos os interessados (webhook) ──
    todos = list(dict.fromkeys(e for e in [vend, chain["tl"], chain["coord"]] + SIGNOFF_CC_FIXO if e))
    html_aceite = (
        f"<p>Olá,</p>"
        f"<p><b>{nome}</b> confirmou e assinou o aceite das comissões de <b>{label}</b> "
        f"em {now_brt}.</p>"
        f"{resumo_html}"
        f"{pdf_link_html}"
        f"<p style='color:#888;font-size:12px;margin-top:16px'>"
        f"Dúvidas? Acesse o <a href='https://commission-dashboard-m7r6fovrsa-uc.a.run.app/dashboard'>dashboard de comissões</a>.</p>"
        f"<p>— Finance · Fluency Academy</p>"
    )
    send_mail(todos, f"Aceite registrado — {nome} · {label}", html_aceite)

    # ── Template B: PDF + XLS em anexo via SMTP (ativo quando FINANCE_SMTP_PASS configurado) ──
    if pdf and FINANCE_SMTP_PASS:
        xls = None
        try:
            xls = _build_signoff_xls(nome, vend, label, cargo, chain["tl"], summary, txs)
        except Exception as _xe:
            app.logger.warning("XLS fechamento falhou: %s", _xe)
        html_pdf = (
            f"<p>Olá, {nome},</p>"
            f"<p>Seu aceite de comissões de <b>{label}</b> foi registrado com sucesso em {now_brt}. "
            f"O demonstrativo detalhado segue em anexo (PDF e XLS).</p>"
            f"{resumo_html}"
            f"<p>Qualquer dúvida, procure seu Team Leader.</p>"
            f"<p>— Finance · Fluency Academy</p>"
        )
        base = f"fechamento_{vend.split('@')[0]}_{mes[:7]}"
        attachments = [(f"{base}.pdf", pdf, "application/pdf")]
        if xls:
            attachments.append((f"{base}.xlsx", xls,
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
        send_mail_smtp([vend], f"Demonstrativo de comissões — {nome} · {label}", html_pdf,
                       attachments=attachments)

    # ── Verifica se o time inteiro do TL assinou → Template C: fechamento completo ──
    try:
        _notify_if_team_complete(chain["tl"], mes, label, chain)
    except Exception as _e:
        app.logger.warning("check fechamento completo falhou: %s", _e)
    # ── Verifica se 100% do roster assinou → aviso People Ops (Milena+Thabita+Matheus) ──
    try:
        _notify_people_ops_if_complete(mes)
    except Exception as _e:
        app.logger.warning("aviso people_ops falhou: %s", _e)
    return jsonify({"ok": True})

def _send_confirmacao_financeiro(vend: str, mes: str, nome: str, label: str,
                                 chain: dict, summary: dict, txs: list):
    """E-mail formal ao financeiro (giullia+paula) com gráfico inline e PDF anexo."""
    if _view_as():
        return
    import base64
    mes_ext = _mes_extenso(mes)
    now_brt = (datetime.now(timezone.utc) - timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")
    chart_png = None
    try:
        dados_chart = {
            "gbv_bruto":  summary.get("gbv_bruto"),
            "gbv_liquido": summary.get("gbv_liq"),
            "gbv_churn":  summary.get("churn"),
            "comissao":   summary.get("vlr_final"),
        }
        chart_png = _render_comissao_chart(dados_chart)
    except Exception as e:
        app.logger.warning("chart geração falhou: %s", e)
    chart_b64 = ("data:image/png;base64," + base64.b64encode(chart_png).decode()) if chart_png else ""
    pct  = f"{summary.get('ating', 0) * 100:.2f}".replace(".", ",")
    mult = f"{summary.get('mult', 0):.1f}".replace(".", ",")
    chart_html = (
        f'<img src="{chart_b64}" alt="GBV e comissão" '
        f'style="max-width:560px;width:100%;height:auto;border:1px solid #E5E5E5;border-radius:8px;" />'
    ) if chart_b64 else (
        f'<table style="border-collapse:collapse;margin:8px 0;font-size:14px;">'
        f'<tr><td style="padding:4px 12px 4px 0;color:#666">GBV bruto</td>'
        f'<td><b>{_brl(summary.get("gbv_bruto"))}</b></td></tr>'
        f'<tr><td style="padding:4px 12px 4px 0;color:#666">GBV líquido</td>'
        f'<td><b>{_brl(summary.get("gbv_liq"))}</b></td></tr>'
        f'<tr><td style="padding:4px 12px 4px 0;color:#666">Churn</td>'
        f'<td><b>{_brl(summary.get("churn"))}</b></td></tr></table>'
    )
    html_body = (
        f"<p>Olá, Giullia e Paula,</p>"
        f"<p>O time financeiro recebeu o <strong>aceite formal de comissão</strong> de:</p>"
        f"<table style='border-collapse:collapse;margin:8px 0;'>"
        f"<tr><td style='padding:4px 12px 4px 0;color:#666'>Vendedor:</td>"
        f"<td><strong>{nome}</strong></td></tr>"
        f"<tr><td style='padding:4px 12px 4px 0;color:#666'>E-mail:</td>"
        f"<td>{vend}</td></tr>"
        f"<tr><td style='padding:4px 12px 4px 0;color:#666'>Competência:</td>"
        f"<td>{mes_ext}</td></tr>"
        f"<tr><td style='padding:4px 12px 4px 0;color:#666'>Aceite em:</td>"
        f"<td>{now_brt}</td></tr></table>"
        f"<p>Números do fechamento:</p>"
        f"{chart_html}"
        f"<p style='background:#FAFAFA;border-left:3px solid #7B5CF6;padding:10px 14px;margin:14px 0;'>"
        f"Comissão apurada: <strong style='color:#7B5CF6;font-size:18px;'>"
        f"{_brl(summary.get('vlr_final'))}</strong></p>"
        f"<p>O demonstrativo completo segue em anexo (PDF). Detalhamento por transação disponível no "
        f"<a href='https://commission-dashboard-302941366897.us-central1.run.app/dashboard'>"
        f"dashboard de comissão</a>.</p>"
        f"<p style='color:#888;font-size:12px;margin-top:18px;'>"
        f"Esta é uma mensagem automática — <strong>não responda a este e-mail</strong>. "
        f"Qualquer dúvida, procure seu Team Leader.</p>"
        f"<p>Atenciosamente,<br><strong>Finance — Fluency Academy</strong><br>"
        f"finance@fluencyacademy.io</p>"
    )
    pdf_bytes = None
    cargo = _load_vmap().get(vend, {}).get("cargo", "")
    try:
        pdf_bytes = _build_signoff_pdf(nome, vend, label, cargo, chain.get("tl"), summary, txs)
    except Exception as e:
        app.logger.warning("PDF geração falhou (confirmacao financeiro): %s", e)
    atts = [(f"comissao-{vend.split('@')[0]}-{mes[:7]}.pdf", pdf_bytes, "application/pdf")] if pdf_bytes else None
    cc = [e for e in (chain.get("tl"), chain.get("coord"))
          if e and e.lower() not in [x.lower() for x in FINANCE_CONFIRMACAO_EMAILS]]
    send_mail_smtp(
        FINANCE_CONFIRMACAO_EMAILS,
        f"Aceite de comissão recebido — {nome} · {mes_ext}",
        html_body, cc=cc, attachments=atts
    )

# ── API: extras (HP que o sistema não trouxe) ─────────────────────────────────

def _resolve_hp(transaction_id: str):
    """Busca um código HP no obt_conversions → (gbv, modality_payment, is_churn) ou None."""
    rows = run_query("""
        SELECT SUM(CAST(gbv AS NUMERIC))     AS gbv,
               ANY_VALUE(modality_payment)   AS modality_payment,
               MAX(CAST(is_churn AS INT64))  AS is_churn
        FROM `fluency-gold.conversion.obt_conversions`
        WHERE CAST(transaction_id AS STRING) = @hp
    """, [bigquery.ScalarQueryParameter("hp", "STRING", transaction_id)])
    if not rows or rows[0].get("gbv") is None:
        return None
    r = rows[0]
    return (float(r["gbv"]), r.get("modality_payment"), int(r["is_churn"] or 0))

@app.route("/api/extras", methods=["GET"])
@login_required
def api_extras_get():
    target = effective_email()
    mes    = resolve_month()
    rows = run_query("""
        SELECT id, transaction_id, gbv, modality_payment, is_churn, fonte, nota,
               status_tl, status_coord, created_at
        FROM `fluency-finance.commission.extras_vendedores`
        WHERE LOWER(vendedor) = LOWER(@v) AND competencia = DATE(@mes)
        ORDER BY created_at DESC
    """, [
        bigquery.ScalarQueryParameter("v",   "STRING", target),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])
    for r in rows:
        r["gbv"] = float(r["gbv"]) if r.get("gbv") is not None else None
        r["is_churn"] = int(r["is_churn"]) if r.get("is_churn") is not None else None
        r["created_at"] = str(r["created_at"]) if r.get("created_at") else None
    return jsonify(rows)

@app.route("/api/extras", methods=["POST"])
@login_required
def api_extras_post():
    if _get_role_data()["role"] in READONLY_ROLES or _view_as():
        return jsonify({"error": "forbidden"}), 403
    vend = session["email"].lower()
    body = request.get_json(silent=True) or {}
    hp   = (body.get("transaction_id") or "").strip()
    nota = (body.get("nota") or "").strip()[:500]
    mes  = body.get("mes")
    mes  = (mes + "-01") if (mes and len(mes) == 7) else current_month_brt()
    if not hp:
        return jsonify({"error": "código HP é obrigatório"}), 400

    # Já está no sistema (commission.vendedores) para este vendedor/mês? Então não é "extra".
    dup_sys = run_query("""
        SELECT 1 FROM `fluency-finance.commission.vendedores`
        WHERE CAST(transaction_id AS STRING) = @hp AND LOWER(vendedor) = LOWER(@v)
          AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes) LIMIT 1
    """, [
        bigquery.ScalarQueryParameter("hp",  "STRING", hp),
        bigquery.ScalarQueryParameter("v",   "STRING", vend),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])
    if dup_sys:
        return jsonify({"error": "Esse HP já está nas suas transações do sistema."}), 409
    # Já adicionado antes?
    dup_ext = run_query("""
        SELECT 1 FROM `fluency-finance.commission.extras_vendedores`
        WHERE transaction_id = @hp AND LOWER(vendedor) = LOWER(@v) AND competencia = DATE(@mes) LIMIT 1
    """, [
        bigquery.ScalarQueryParameter("hp",  "STRING", hp),
        bigquery.ScalarQueryParameter("v",   "STRING", vend),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])
    if dup_ext:
        return jsonify({"error": "Você já adicionou esse HP neste mês."}), 409

    # Bloqueia inclusão de HP após aceite do comissionamento
    signed = run_query("""
        SELECT 1 FROM `fluency-finance.commission.signoff_vendedores`
        WHERE LOWER(vendedor) = LOWER(@v) AND competencia = DATE(@mes) LIMIT 1
    """, [
        bigquery.ScalarQueryParameter("v",   "STRING", vend),
        bigquery.ScalarQueryParameter("mes", "DATE",   mes),
    ])
    if signed and not _is_newcomer_assist(vend, mes):
        return jsonify({"error": "Não é possível adicionar HP após o envio do aceite de comissionamento."}), 409

    resolved = _resolve_hp(hp)
    gbv, modalidade, is_churn = resolved if resolved else (None, None, None)
    fonte = "obt" if resolved else "nao_localizado"
    chain = _commission_chain(vend, mes)
    eid = hashlib.sha1(f"{vend}|{hp}|{datetime.now(timezone.utc).timestamp()}".encode()).hexdigest()[:16]

    bq.query("""
        INSERT INTO `fluency-finance.commission.extras_vendedores`
          (id, transaction_id, vendedor, competencia, gbv, modality_payment, is_churn,
           fonte, nota, created_at, created_by, status_tl, tl_email, status_coord, coord_email)
        VALUES
          (@id, @hp, @v, DATE(@mes), @gbv, @mod, @churn, @fonte, @nota,
           CURRENT_TIMESTAMP(), @v, 'pendente', @tl, 'pendente', @coord)
    """, job_config=bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("id",    "STRING",  eid),
        bigquery.ScalarQueryParameter("hp",    "STRING",  hp),
        bigquery.ScalarQueryParameter("v",     "STRING",  vend),
        bigquery.ScalarQueryParameter("mes",   "DATE",    mes),
        bigquery.ScalarQueryParameter("gbv",   "NUMERIC", gbv),
        bigquery.ScalarQueryParameter("mod",   "STRING",  modalidade),
        bigquery.ScalarQueryParameter("churn", "INT64",   is_churn),
        bigquery.ScalarQueryParameter("fonte", "STRING",  fonte),
        bigquery.ScalarQueryParameter("nota",  "STRING",  nota),
        bigquery.ScalarQueryParameter("tl",    "STRING",  chain["tl"]),
        bigquery.ScalarQueryParameter("coord", "STRING",  chain["coord"]),
    ])).result()

    nome = _name_from_email(vend); label = _mes_label(mes)
    valor = f"R$ {gbv:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if gbv is not None else "valor não localizado no sistema"
    html = (f"<p>Olá,</p><p><b>{nome}</b> adicionou um HP que não estava no sistema, para o fechamento de "
            f"<b>{label}</b>, e ele aguarda aprovação:</p>"
            f"<ul><li><b>HP:</b> {hp}</li><li><b>GBV:</b> {valor}</li>"
            f"<li><b>Modalidade:</b> {modalidade or '—'}</li><li><b>Obs.:</b> {nota or '—'}</li></ul>"
            f"<p>Aprove ou rejeite na aba <b>Aprovações</b> do dashboard. "
            f"O valor só entra no número do vendedor quando TL e coordenador aprovarem.</p>")
    send_mail([chain["tl"], chain["coord"]], f"Novo HP p/ aprovação — {nome} ({label})", html, cc=[chain["master"]])
    return jsonify({"ok": True, "id": eid, "fonte": fonte, "gbv": gbv})

@app.route("/api/extras/pending", methods=["GET"])
@login_required
def api_extras_pending():
    role = _get_role_data()["role"]
    me   = session["email"].lower()
    if is_master():         # master: tudo pendente (qualquer lado)
        where = "status_tl = 'pendente' OR status_coord = 'pendente'"
        params = []
    elif role == "gestor":  # coordenador: tudo aguardando aprovação do coord
        where = "status_coord = 'pendente'"
        params = []
    elif role == "tl":      # TL: extras do MEU time aguardando minha aprovação
        where = "status_tl = 'pendente' AND LOWER(tl_email) = LOWER(@me)"
        params = [bigquery.ScalarQueryParameter("me", "STRING", me)]
    else:
        return jsonify([])
    rows = run_query(f"""
        SELECT id, transaction_id, vendedor, competencia, gbv, modality_payment, is_churn,
               fonte, nota, status_tl, status_coord, tl_email, coord_email, created_at
        FROM `fluency-finance.commission.extras_vendedores`
        WHERE {where}
        ORDER BY created_at DESC
    """, params)
    for r in rows:
        r["gbv"] = float(r["gbv"]) if r.get("gbv") is not None else None
        r["is_churn"] = int(r["is_churn"]) if r.get("is_churn") is not None else None
        r["competencia"] = str(r["competencia"]) if r.get("competencia") else None
        r["created_at"] = str(r["created_at"]) if r.get("created_at") else None
        r["vendedor_nome"] = _name_from_email(r["vendedor"])
    return jsonify(rows)

@app.route("/api/extras/<eid>/decision", methods=["POST"])
@login_required
def api_extras_decision(eid):
    role = _get_role_data()["role"]
    if role not in ("tl", "gestor") or _view_as():
        return jsonify({"error": "forbidden"}), 403
    me   = session["email"].lower()
    body = request.get_json(silent=True) or {}
    decision = body.get("decision")
    if decision not in ("aprovado", "rejeitado"):
        return jsonify({"error": "decision inválida"}), 400

    rows = run_query("""
        SELECT id, vendedor, competencia, gbv, transaction_id, tl_email, coord_email,
               status_tl, status_coord
        FROM `fluency-finance.commission.extras_vendedores` WHERE id = @id LIMIT 1
    """, [bigquery.ScalarQueryParameter("id", "STRING", eid)])
    if not rows:
        return jsonify({"error": "extra não encontrado"}), 404
    ex = rows[0]

    sets, params = [], [bigquery.ScalarQueryParameter("id", "STRING", eid),
                        bigquery.ScalarQueryParameter("dec", "STRING", decision),
                        bigquery.ScalarQueryParameter("me", "STRING", me)]
    is_coord = role == "gestor"
    is_tl    = (ex.get("tl_email") or "").lower() == me
    if is_coord:
        sets.append("status_coord = @dec, coord_email = @me, coord_at = CURRENT_TIMESTAMP()")
    if is_tl:   # TL do time (ou coordenador que também é o TL direto) marca o lado TL
        sets.append("status_tl = @dec, tl_email = @me, tl_at = CURRENT_TIMESTAMP()")
    if not sets:
        return jsonify({"error": "fora do seu escopo"}), 403
    bq.query(f"UPDATE `fluency-finance.commission.extras_vendedores` SET {', '.join(sets)} WHERE id = @id",
             job_config=bigquery.QueryJobConfig(query_parameters=params)).result()

    # Estado final após update
    new_tl    = decision if is_tl else ex.get("status_tl")
    new_coord = decision if is_coord else ex.get("status_coord")
    if new_tl == "aprovado" and new_coord == "aprovado":
        vend = ex["vendedor"].lower(); mes = str(ex["competencia"])
        chain = _commission_chain(vend, mes); label = _mes_label(mes)
        html = (f"<p>Olá,</p><p>O HP <b>{ex.get('transaction_id')}</b> que você adicionou para "
                f"<b>{label}</b> foi <b>aprovado por TL e coordenador</b> e já está somado ao seu "
                f"GBV no dashboard.</p>")
        send_mail([vend], f"HP aprovado e incorporado — {label}", html, cc=[chain["master"]])
    return jsonify({"ok": True, "status_tl": new_tl, "status_coord": new_coord})

# ── API: receivables ─────────────────────────────────────────────────────────

@app.route("/api/receivables")
@login_required
def api_receivables():
    target = resolve_target(effective_email())
    mes    = resolve_month()
    sql = """
        SELECT
          DATE_TRUNC(DATE(COALESCE(
            transaction_confirmation_purchase_at_brt_timestamp,
            contract_created_at_brt_timestamp
          )), MONTH)                              AS mes_pagamento,
          transaction_status,
          COUNT(*)                               AS parcelas,
          CAST(SUM(transaction_amount) AS NUMERIC) AS valor
        FROM `fluency-finance.commission.vendedores`
        WHERE LOWER(vendedor) = LOWER(@email)
          AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
        GROUP BY 1, 2
        ORDER BY 1, 2
    """
    rows = run_query(sql, [
        bigquery.ScalarQueryParameter("email", "STRING", target),
        bigquery.ScalarQueryParameter("mes",   "DATE",   mes),
    ])
    result = []
    for r in rows:
        result.append({
            "mes_pagamento": str(r["mes_pagamento"]) if r["mes_pagamento"] else None,
            "status":        r["transaction_status"],
            "parcelas":      int(r["parcelas"]),
            "valor":         float(r["valor"]),
        })
    return jsonify(result)

# ── Drive / Sheets helpers ────────────────────────────────────────────────────

def _get_drive_sheets():
    """Returns (drive_service, sheets_service) using ambient credentials."""
    from googleapiclient.discovery import build
    creds, _ = _gauth_default(scopes=[
        "https://www.googleapis.com/auth/drive.readonly",
        "https://www.googleapis.com/auth/spreadsheets.readonly",
    ])
    drive_svc  = build("drive",  "v3", credentials=creds)
    sheets_svc = build("sheets", "v4", credentials=creds)
    return drive_svc, sheets_svc


def _find_sales_commission_file(drive_svc, mes: str) -> str | None:
    """mes = 'YYYY-MM'.  Returns Drive file ID or None."""
    month_num = int(mes[5:7])
    year_2    = mes[2:4]
    abbr      = _MONTH_ABBR[month_num - 1]
    target    = f"Sales Commission - {abbr}.{year_2}"
    result    = drive_svc.files().list(
        q=f"'{DRIVE_FOLDER_ID}' in parents and name contains '{target}'",
        fields="files(id,name)",
        pageSize=10,
    ).execute()
    files = [f for f in result.get("files", []) if not f["name"].startswith("Copy")]
    return files[0]["id"] if files else None




def _normalize_status(raw) -> str | None:
    s = str(raw or "").strip().lower()
    if s in _PAID_STATUSES:    return "paid"
    if s in _REFUND_STATUSES:  return "refunded"
    if s in _CHARGE_STATUSES:  return "chargeback"
    return None


def _safe_numeric(v) -> float:
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _parse_date(v) -> str | None:
    """Converts M/D/YYYY or M/D/YYYY H:MM:SS to YYYY-MM-DD, or returns ISO string."""
    s = str(v or "").strip()
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None


def _read_vendor_rows(sheets_svc, file_id: str, mes: str) -> list[dict]:
    """
    Reads all vendor tabs from the Sales Commission file.
    Returns a list of BQ-ready dicts for the comissao_apurada table.
    """
    spreadsheet = sheets_svc.spreadsheets().get(spreadsheetId=file_id).execute()
    sheet_names = [s["properties"]["title"] for s in spreadsheet.get("sheets", [])]

    competencia = mes + "-01"
    all_rows = []

    for sheet_name in sheet_names:
        raw = sheets_svc.spreadsheets().values().get(
            spreadsheetId=file_id,
            range=f"'{sheet_name}'!A:J",
            valueRenderOption="UNFORMATTED_VALUE",
        ).execute().get("values", [])

        if not raw:
            continue

        # Locate the data header row (has 'buyer_email' or 'sales_rep_clean')
        header_idx = None
        for i, row in enumerate(raw):
            lower_row = [str(c).lower().strip() for c in row]
            if "buyer_email" in lower_row or "sales_rep_clean" in lower_row:
                header_idx = i
                break
        if header_idx is None:
            continue

        for row in raw[header_idx + 1:]:
            if len(row) < 5:
                continue
            # Stop at metadata footer ("Colaborador" row or filler date 3799)
            if str(row[0]).strip().lower() in ("colaborador", ""):
                if len(row) < 3 or not row[1]:
                    break
                continue
            date_val = _parse_date(row[9]) if len(row) > 9 else _parse_date(row[2])
            if date_val and date_val.startswith("3799"):
                continue

            status = _normalize_status(row[6] if len(row) > 6 else "")
            if status is None:
                continue

            all_rows.append({
                "sales_rep":   str(row[0]).strip(),
                "buyer_email": str(row[1]).strip() if len(row) > 1 else "",
                "data_compra": _parse_date(row[2]) if len(row) > 2 else None,
                "tipo_pagto":  str(row[3]).strip() if len(row) > 3 else "",
                "gbv":         _safe_numeric(row[4]) if len(row) > 4 else 0.0,
                "comissao":    _safe_numeric(row[5]) if len(row) > 5 else 0.0,
                "status":      status,
                "canal_vendas":str(row[7]).strip() if len(row) > 7 else "",
                "week_label":  str(row[8]).strip() if len(row) > 8 else "",
                "competencia": competencia,
                "arquivo_id":  file_id,
            })

    return all_rows


def _import_drive_rows_to_bq(rows: list[dict], mes: str) -> int:
    """Inserts rows into comissao_apurada, replacing existing data for the month."""
    competencia = mes + "-01"

    # Delete existing rows for this month
    run_query(
        "DELETE FROM `fluency-finance.commission.comissao_apurada` WHERE competencia = DATE(@mes)",
        [bigquery.ScalarQueryParameter("mes", "DATE", competencia)],
    )

    if not rows:
        return 0

    # Batch insert via BQ load (streaming)
    table_ref = bq.dataset("commission").table("comissao_apurada")
    bq_rows = []
    for r in rows:
        bq_rows.append({
            "sales_rep":   r["sales_rep"],
            "buyer_email": r["buyer_email"],
            "data_compra": r["data_compra"],
            "tipo_pagto":  r["tipo_pagto"],
            "gbv":         str(r["gbv"]),
            "comissao":    str(r["comissao"]),
            "status":      r["status"],
            "canal_vendas":r["canal_vendas"],
            "week_label":  r["week_label"],
            "competencia": competencia,
            "arquivo_id":  r["arquivo_id"],
            "importado_em": datetime.now(timezone.utc).isoformat(),
        })

    errors = bq.insert_rows_json(table_ref, bq_rows)
    if errors:
        raise RuntimeError(f"BQ insert errors: {errors[:3]}")
    return len(bq_rows)


def _get_tl_supplement_metas(sheets_svc, mes: str) -> dict:
    """
    Reads the vendor-config section of the Meta tab (SHEETS_META_ID) to get the
    vendor→TL mapping (col[0]=vendor email, col[6]=TL email).  For every TL that
    has no direct row in meta_vendedores for the given month, computes their meta
    as the sum of their vendors' metas from BQ.
    mes = 'YYYY-MM'.  Returns {lowercase_tl_email: float_meta}.
    """
    try:
        resp = sheets_svc.spreadsheets().values().get(
            spreadsheetId=SHEETS_META_ID,
            range="Meta",
            valueRenderOption="FORMATTED_VALUE",
        ).execute()
    except Exception:
        return {}

    vendor_to_tl: dict[str, str] = {}
    for row in resp.get("values", []):
        if len(row) < 3:
            continue
        vendor = str(row[0]).strip().lower()
        tl     = str(row[2]).strip().lower()   # col 2 = "Gestor"
        if "@" in vendor and "@" in tl and "#n/a" not in tl:
            vendor_to_tl[vendor] = tl

    if not vendor_to_tl:
        return {}

    all_tls     = set(vendor_to_tl.values())
    competencia = mes + "-01"

    # Only treat TLs with a positive meta as "already covered" —
    # TLs with 0 or NULL meta in the sheet also need the vendor-sum treatment.
    existing = {
        r["vkey"]
        for r in run_query(
            "SELECT LOWER(email_vendedor) AS vkey "
            "FROM `fluency-finance.commission.hierarquia_comercial` "
            "WHERE mes_venda = DATE(@mes) AND COALESCE(valor_meta, 0) > 0",
            [bigquery.ScalarQueryParameter("mes", "DATE", competencia)],
        )
    }

    tl_metas: dict[str, float] = {}
    for tl_email in all_tls - existing:
        vendors = [v for v, t in vendor_to_tl.items() if t == tl_email]
        if not vendors:
            continue
        safe_emails = [e.replace("'", "") for e in vendors]
        in_list     = ", ".join(f"'{e}'" for e in safe_emails)
        rows_bq = run_query(
            f"SELECT COALESCE(SUM(valor_meta), 0) AS total "
            f"FROM `fluency-finance.commission.hierarquia_comercial` "
            f"WHERE mes_venda = DATE(@mes) "
            f"AND LOWER(email_vendedor) IN ({in_list})",
            [bigquery.ScalarQueryParameter("mes", "DATE", competencia)],
        )
        total = float(rows_bq[0]["total"]) if rows_bq else 0.0
        if total > 0:
            tl_metas[tl_email] = total

    return tl_metas


def _fix_tl_metas_in_historica(mes: str, sheets_svc) -> int:
    """
    After INSERT from vw_comissao, TL rows land with NULL multiplicador because
    their meta_vendedores entry is 0 or NULL.  This function:
      1. Calls _get_tl_supplement_metas to get {tl_email: meta_float} (already
         handles TLs absent *or* with 0/NULL meta via the updated existing-check).
      2. For each TL with a positive computed meta, UPDATEs comissao_historica
         with the correct atingimento_meta / multiplicador / vlr_final_comissao.
    Returns the number of TL rows updated.
    """
    tl_metas = _get_tl_supplement_metas(sheets_svc, mes)
    if not tl_metas:
        return 0

    competencia = mes + "-01"
    updated = 0
    for tl_email, meta in tl_metas.items():
        if meta <= 0:
            continue
        run_query("""
            UPDATE `fluency-finance.commission.comissao_historica`
            SET
              atingimento_meta   = SAFE_DIVIDE(
                CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta
              ),
              multiplicador      = CASE
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 0.75 THEN 0.3
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 0.98 THEN 0.5
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 1.20 THEN 1.0
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 1.30 THEN 1.2
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 1.50 THEN 1.3
                ELSE 1.5
              END,
              vlr_final_comissao = total_comissao * CASE
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 0.75 THEN 0.3
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 0.98 THEN 0.5
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 1.20 THEN 1.0
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 1.30 THEN 1.2
                WHEN SAFE_DIVIDE(CAST(gbv_churn_descontado_transaction AS FLOAT64), @meta) < 1.50 THEN 1.3
                ELSE 1.5
              END
            WHERE LOWER(vendedor) = LOWER(@email)
              AND DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
        """, [
            bigquery.ScalarQueryParameter("meta",  "FLOAT64", meta),
            bigquery.ScalarQueryParameter("email", "STRING",  tl_email),
            bigquery.ScalarQueryParameter("mes",   "DATE",    competencia),
        ])
        updated += 1
    return updated


def _rebuild_comissao_historica_from_apurada(mes: str, sheets_svc=None):
    """
    Aggregates comissao_apurada → comissao_historica for the given month,
    joining meta_vendedores for multiplicador/atingimento_meta.
    If sheets_svc is provided, also computes metas for TLs absent from
    meta_vendedores (e.g. alexandre.kim) by summing their vendors' metas.
    Falls back to multiplicador=0 if meta table is inaccessible.
    """
    competencia = mes + "-01"

    # Build supplemental TL metas (for TLs absent from meta_vendedores BQ)
    tl_metas: dict[str, float] = {}
    if sheets_svc is not None:
        try:
            tl_metas = _get_tl_supplement_metas(sheets_svc, mes)
        except Exception:
            pass  # Non-fatal

    if tl_metas:
        structs   = ", ".join(
            f"STRUCT('{e.replace(chr(39), '')}' AS email, {m:.2f} AS meta)"
            for e, m in tl_metas.items()
        )
        tl_union  = (
            "\n        UNION ALL\n"
            f"        SELECT LOWER(s.email) AS vkey, s.meta AS valor_meta\n"
            f"        FROM UNNEST([{structs}]) AS s\n"
        )
    else:
        tl_union = ""

    run_query(
        "DELETE FROM `fluency-finance.commission.comissao_historica` "
        "WHERE DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)",
        [bigquery.ScalarQueryParameter("mes", "DATE", competencia)],
    )
    run_query(
        f"""
        INSERT INTO `fluency-finance.commission.comissao_historica`
          (vendedor, contract_created_at_brt_timestamp,
           transaction_confirmation_purchase_at_brt_timestamp,
           gbv, qtd_is_churn_transaction,
           gbv_apenas_churn_transaction, gbv_churn_descontado_transaction,
           comissao_inteligente, comissao_parcelado, comissao_a_vista,
           atingimento_meta, multiplicador, total_comissao, vlr_final_comissao)
        WITH agg AS (
          SELECT
            sales_rep                                                    AS vendedor,
            DATETIME(DATE(@mes))                                         AS ts,
            SUM(gbv)                                                     AS gbv,
            SUM(CASE WHEN status = 'paid' THEN gbv ELSE 0 END)          AS gbv_liq,
            SUM(CASE WHEN status IN ('refunded','chargeback')
                     THEN gbv ELSE 0 END)                                AS gbv_churn,
            COUNT(CASE WHEN status IN ('refunded','chargeback')
                       THEN 1 END)                                       AS qtd_churn,
            SUM(CASE WHEN LOWER(tipo_pagto) LIKE '%intelig%'
                      AND status = 'paid' THEN comissao ELSE 0 END)     AS com_int,
            SUM(CASE WHEN LOWER(tipo_pagto) LIKE '%parcel%'
                      AND status = 'paid' THEN comissao ELSE 0 END)     AS com_parc,
            SUM(CASE WHEN (LOWER(tipo_pagto) LIKE '%vista%'
                        OR LOWER(tipo_pagto) LIKE '%recorr%')
                      AND status = 'paid' THEN comissao ELSE 0 END)     AS com_av,
            SUM(CASE WHEN status = 'paid' THEN comissao ELSE 0 END)     AS com_total
          FROM `fluency-finance.commission.comissao_apurada`
          WHERE competencia = DATE(@mes)
          GROUP BY 1, 2
        ),
        meta AS (
          SELECT LOWER(email_vendedor) AS vkey, valor_meta
          FROM `fluency-finance.commission.hierarquia_comercial`
          WHERE mes_venda = DATE(@mes){tl_union}
        ),
        calc AS (
          SELECT
            a.vendedor,
            a.ts,
            a.gbv,
            a.qtd_churn,
            a.gbv_churn,
            a.gbv_liq,
            a.com_int,
            a.com_parc,
            a.com_av,
            a.com_total,
            CASE WHEN COALESCE(m.valor_meta, 0) = 0 THEN 0.0
                 ELSE SAFE_DIVIDE(a.gbv_liq, m.valor_meta)
            END AS ating,
            CASE WHEN COALESCE(m.valor_meta, 0) = 0 THEN 0.0
                 WHEN SAFE_DIVIDE(a.gbv_liq, m.valor_meta) < 0.75 THEN 0.3
                 WHEN SAFE_DIVIDE(a.gbv_liq, m.valor_meta) < 0.98 THEN 0.5
                 WHEN SAFE_DIVIDE(a.gbv_liq, m.valor_meta) < 1.20 THEN 1.0
                 WHEN SAFE_DIVIDE(a.gbv_liq, m.valor_meta) < 1.30 THEN 1.2
                 WHEN SAFE_DIVIDE(a.gbv_liq, m.valor_meta) < 1.50 THEN 1.3
                 ELSE 1.5
            END AS mult
          FROM agg a
          LEFT JOIN meta m ON LOWER(a.vendedor) = m.vkey
        )
        SELECT
          vendedor,
          ts,
          NULL,
          gbv,
          qtd_churn,
          gbv_churn,
          gbv_liq,
          com_int,
          com_parc,
          com_av,
          ating,
          mult,
          com_total,
          com_total * mult
        FROM calc
        """,
        [bigquery.ScalarQueryParameter("mes", "DATE", competencia)],
    )


# ── Admin: refresh automático ─────────────────────────────────────────────────

@app.route("/admin/refresh-current", methods=["POST"])
@require_refresh_secret
def admin_refresh_current():
    mes = current_month_brt()
    run_query("""
        DELETE FROM `fluency-finance.commission.comissao_historica`
        WHERE DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)])
    run_query("""
        INSERT INTO `fluency-finance.commission.comissao_historica`
        SELECT * FROM `fluency-finance.commission.vw_comissao`
        WHERE DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)
    """, [bigquery.ScalarQueryParameter("mes", "DATE", mes)])
    try:
        _, sheets_svc = _get_drive_sheets()
        n_tls = _fix_tl_metas_in_historica(mes[:7], sheets_svc)
    except Exception:
        n_tls = 0
    return jsonify({"ok": True, "mes_atualizado": mes, "tls_corrigidos": n_tls})

@app.route("/admin/import-drive", methods=["POST"])
@require_refresh_secret
def admin_import_drive():
    """
    Imports the audited "Sales Commission" Drive file for a given month into
    comissao_apurada, then rebuilds comissao_historica from it.

    Body JSON: {"mes": "YYYY-MM"}   (defaults to previous month if omitted)
    """
    body = request.get_json(silent=True) or {}
    mes  = body.get("mes", "").strip()  # expects "YYYY-MM"
    if not mes or not re.match(r"^\d{4}-\d{2}$", mes):
        # Default to previous month
        brt = datetime.now(timezone.utc) - timedelta(hours=3)
        if brt.month == 1:
            mes = f"{brt.year - 1}-12"
        else:
            mes = f"{brt.year}-{brt.month - 1:02d}"

    try:
        drive_svc, sheets_svc = _get_drive_sheets()
    except Exception as e:
        return jsonify({"error": f"Falha ao inicializar clientes Drive/Sheets: {e}"}), 500

    file_id = body.get("file_id", "").strip() or _find_sales_commission_file(drive_svc, mes)
    if not file_id:
        return jsonify({"error": f"Arquivo 'Sales Commission' não encontrado para {mes}"}), 404

    try:
        rows = _read_vendor_rows(sheets_svc, file_id, mes)
    except Exception as e:
        return jsonify({"error": f"Falha ao ler planilha: {e}"}), 500

    if not rows:
        return jsonify({"error": f"Nenhuma linha de dados encontrada na planilha de {mes}"}), 400

    try:
        count = _import_drive_rows_to_bq(rows, mes)
    except Exception as e:
        return jsonify({"error": f"Falha ao gravar comissao_apurada: {e}"}), 500

    try:
        _rebuild_comissao_historica_from_apurada(mes, sheets_svc=sheets_svc)
    except Exception as e:
        return jsonify({
            "ok": True,
            "mes": mes,
            "linhas_importadas": count,
            "aviso": f"Linhas importadas para comissao_apurada, mas recálculo de comissao_historica falhou: {e}. "
                     "Verifique se meta_vendedores está acessível.",
        }), 207

    vendedores = sorted({r["sales_rep"] for r in rows})
    return jsonify({
        "ok":               True,
        "mes":              mes,
        "arquivo_id":       file_id,
        "linhas_importadas": count,
        "vendedores":       vendedores,
    })


@app.route("/admin/close-previous", methods=["POST"])
@require_refresh_secret
def admin_close_previous():
    """
    Closes previous month from BQ (vw_comissao).
    Drive import is available separately via /admin/import-drive for historical months.
    """
    brt = datetime.now(timezone.utc) - timedelta(hours=3)
    if brt.month == 1:
        prev = f"{brt.year - 1}-12-01"
    else:
        prev = f"{brt.year}-{brt.month - 1:02d}-01"

    run_query(
        "DELETE FROM `fluency-finance.commission.comissao_historica` "
        "WHERE DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)",
        [bigquery.ScalarQueryParameter("mes", "DATE", prev)],
    )
    run_query(
        "INSERT INTO `fluency-finance.commission.comissao_historica` "
        "SELECT * FROM `fluency-finance.commission.vw_comissao` "
        "WHERE DATE_TRUNC(DATE(contract_created_at_brt_timestamp), MONTH) = DATE(@mes)",
        [bigquery.ScalarQueryParameter("mes", "DATE", prev)],
    )
    try:
        _, sheets_svc = _get_drive_sheets()
        n_tls = _fix_tl_metas_in_historica(prev[:7], sheets_svc)
    except Exception:
        n_tls = 0
    return jsonify({"ok": True, "mes_fechado": prev, "fonte": "vw_comissao", "tls_corrigidos": n_tls})

# ── API: test-email ──────────────────────────────────────────────────────────

_TEST_EMAIL_TO = [
    "milena.custodio@fluencyacademy.io",
    "liliane.noga@fluencyacademy.io",
    "raphael.bride@fluencyacademy.io",
]

@app.route("/api/test-email", methods=["POST"])
@login_required
def api_test_email():
    if not _real_is_master():
        return jsonify({"error": "forbidden"}), 403
    tipo = request.args.get("tipo", "all")
    sent = []

    _nome = "Vendedor Teste"
    _mes_ext = "Junho/2026"
    _summary = {"n": 12, "gbv_bruto": 85000.0, "gbv_liq": 78000.0, "churn": 7000.0,
                "ating": 1.05, "mult": 1.2, "vlr_final": 3200.0}

    if tipo in ("fechamento", "all"):
        # 1. E-mail ao vendedor (aceite de comissão) — SMTP
        html_vend = (
            "<p>[TESTE — não é real]</p>"
            f"<p>Recebemos neste momento o <b>aceite do colaborador {_nome}</b>, "
            f"confirmando a <b>conferência e concordância</b> com as comissões de <b>{_mes_ext}</b>.</p>"
            "<p>Esta confirmação foi registrada e está <b>salva em nosso banco de dados</b>.</p>"
            "<p>— Fluency · Finance</p>"
        )
        send_mail_smtp(_TEST_EMAIL_TO,
                       f"[TESTE] Confirmação de conferência de comissões — {_nome} — {_mes_ext}",
                       html_vend)
        sent.append("fechamento_vendedor_smtp")

        # 2. E-mail ao financeiro (com gráfico) — SMTP
        import base64
        chart_png = None
        try:
            chart_png = _render_comissao_chart({
                "gbv_bruto": _summary["gbv_bruto"], "gbv_liquido": _summary["gbv_liq"],
                "gbv_churn": _summary["churn"],     "comissao":    _summary["vlr_final"],
            })
        except Exception as _ce:
            app.logger.warning("test-email chart: %s", _ce)
        chart_b64 = ("data:image/png;base64," + base64.b64encode(chart_png).decode()) if chart_png else ""
        chart_html = (
            f'<img src="{chart_b64}" alt="GBV e comissão" '
            f'style="max-width:560px;width:100%;height:auto;border:1px solid #E5E5E5;border-radius:8px;"/>'
        ) if chart_b64 else "<p>[gráfico indisponível]</p>"
        html_fin = (
            "<p>[TESTE — não é real]</p>"
            "<p>Olá, equipe Finance,</p>"
            f"<p>O time financeiro recebeu o <strong>aceite formal de comissão</strong> de "
            f"<strong>{_nome}</strong> para <strong>{_mes_ext}</strong>.</p>"
            f"{chart_html}"
            f"<p style='background:#FAFAFA;border-left:3px solid #7B5CF6;padding:10px 14px;'>"
            f"Comissão apurada: <strong style='color:#7B5CF6;font-size:18px;'>{_brl(_summary['vlr_final'])}</strong></p>"
            "<p>O demonstrativo completo seguiria em anexo (PDF) no e-mail real.</p>"
            "<p>Atenciosamente,<br><strong>Finance — Fluency Academy</strong></p>"
        )
        send_mail_smtp(_TEST_EMAIL_TO,
                       f"[TESTE] Aceite de comissão recebido — {_nome} · {_mes_ext}",
                       html_fin)
        sent.append("fechamento_financeiro_smtp")

    if tipo in ("hp_aprovado", "all"):
        # 3. E-mail ao vendedor quando HP é aprovado — Apps Script webhook
        html_hp = (
            "<p>[TESTE — não é real]</p>"
            "<p>Olá,</p>"
            f"<p>O HP <b>TX_TESTE_001</b> que você adicionou para <b>{_mes_ext}</b> foi "
            "<b>aprovado por TL e coordenador</b> e já está somado ao seu GBV no dashboard.</p>"
        )
        ok_hp = send_mail(_TEST_EMAIL_TO,
                          f"[TESTE] HP aprovado e incorporado — {_mes_ext}",
                          html_hp)
        sent.append(f"hp_aprovado_webhook={'ok' if ok_hp else 'FALHOU'}")

    return jsonify({"ok": True, "sent": sent})

# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)), debug=False)
